# -*- coding: utf-8 -*-
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from pathlib import Path
import openpyxl

path = Path(r'C:\Users\한나\OneDrive\월부\202605 실전 송파구\(송파구) 매임 체크리스트.xlsx')
wb = openpyxl.load_workbook(path)
ws = wb.active

# 이전에 C(물건1)~J(물건8) 입력 완료
# 이번에 K(물건9)~P(물건14) 입력
# 매매가: 만원 단위 → 억 단위 변환 (230,000만원 = 23억)

# (동호수, 매매가(억), 아파트명, 타입, 평형, 층, 전세가(억), 특이사항, 거주자)
properties = [
    ('102동',  23.0,  '위례중앙푸르지오', '111D-1', 33.70, '중/19', 11.0, '어사선 좋음 하동5억', '세입자'),
    ('102동',  24.0,  '위례아이파크',     '131A',   39.70, '고/23', 21.0, '거실2 경사뷰 성실', '세입자'),
    ('101동',  24.0,  '송파와이즈더샵',   '126B1',  38.30, '고/24', 21.0, '거실2 방3', '세입자'),
    ('204동',  23.0,  '아이파크2차',      '118C',   35.80, '5/29',   5.0, '욕방함X', '공실'),
    ('103동',  23.5,  '힐스테이트송파',   '130D',   39.50, '22/29', 22.0, '애리면 인방안기 포함', '세입자'),
    ('104동',  23.5,  '송파푸르지오',     '1,400',  42.40, '15/28', 15.0, '방4 수납9A', '세입자'),
]

start_col = 11  # K열 = 물건9

for i, (dong, price, apt, apt_type, pyeong, floor, jeonse, memo, resident) in enumerate(properties):
    col = start_col + i
    ws.cell(row=3,  column=col).value = apt       # 아파트명
    ws.cell(row=5,  column=col).value = '잠실정원공인'  # 부동산
    ws.cell(row=10, column=col).value = pyeong    # 평형
    ws.cell(row=11, column=col).value = dong      # 동
    ws.cell(row=13, column=col).value = apt_type  # 타입
    ws.cell(row=14, column=col).value = floor     # 층(방향)
    ws.cell(row=15, column=col).value = resident  # 거주자
    ws.cell(row=18, column=col).value = memo      # 특이사항
    ws.cell(row=19, column=col).value = price     # 매매가
    ws.cell(row=21, column=col).value = jeonse    # 전세가

    col_letter = ws.cell(row=1, column=col).column_letter
    print(f"  {col_letter}열 ({apt} {dong}): 매매가 {price}억, 전세가 {jeonse}억 입력 완료")

wb.save(path)
print("\n저장 완료!")
