"""
천안동남구 시트:
- I열(타입) → 첫 숫자 → J열(공급면적) 비어있으면 채우기
- J열(공급면적) → 평형 계산(round(J/3.3058)) → L열 비어있으면 채우기
"""
import sys, io, os, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

BASE_DIR = r'C:\Users\한나\OneDrive\강성업무용\바탕 화면\문서\CLAUDE'
src = os.path.join(BASE_DIR, '시세시트_링크추가.xlsx')

wb = openpyxl.load_workbook(src)
ws = wb['천안동남구']

PYONG_DIVISOR = 3.3058
filled_j = 0
filled_l = 0
skipped = 0

for r in range(5, ws.max_row + 1):
    i_cell = ws.cell(r, 9)   # I 타입
    j_cell = ws.cell(r, 10)  # J 공급면적
    l_cell = ws.cell(r, 12)  # L 평형

    # J열 채우기
    if j_cell.value is None and i_cell.value is not None:
        s = str(i_cell.value)
        m = re.search(r'\d+', s)
        if m:
            j_cell.value = int(m.group())
            filled_j += 1

    # L열 채우기 (J 갱신 후 값 기준)
    if l_cell.value is None and j_cell.value is not None:
        try:
            pyeong = round(int(j_cell.value) / PYONG_DIVISOR)
            l_cell.value = pyeong
            filled_l += 1
        except Exception:
            skipped += 1

wb.save(src)
print(f'공급면적(J) 채움: {filled_j}개')
print(f'평형(L) 채움: {filled_l}개')
print(f'스킵: {skipped}개')
print(f'저장: {src}')
