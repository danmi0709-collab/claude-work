"""
남구.xlsx(부산 남구 단지 목록)의 데이터를 시세시트_링크추가.xlsx의 '부산 남구' 탭에 채움.
- 컬럼 매핑:
   남구.C(읍면동)   → A(동)
   남구.E(단지코드) → D(단지코드)
   남구.F(단지명)   → E(단지명) + 하이퍼링크
   남구.G(사용승인일)→ F
   남구.H(총세대수) → G
   (나이 수식)      → H
   남구.M(평형명=공급)→ J(공급면적)
   남구.I(전용면적) → K
   (평형 계산)      → L
   남구.J(구조)     → M
   남구.K(방수)     → N
   남구.L(욕실수)   → O
   남구.N(평형별세대수)→ P
- 단지명 하이퍼링크: new.land.naver.com/complexes/{단지코드}?ms=...
"""
import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from openpyxl.styles import Font
from copy import copy

BASE = r'C:\Users\한나\OneDrive\강성업무용\바탕 화면\문서\CLAUDE'
SRC = os.path.join(BASE, '남구.xlsx')
DST = os.path.join(BASE, '시세시트_링크추가.xlsx')
URL = 'https://new.land.naver.com/complexes/{}?ms=2AfAQG,3zprE6,17&a=APT:ABYG:JGC&b=A1:B1&e=RETAIL&h=34&i=132&l=198&ad=true'
PYONG_DIVISOR = 3.3058

# 남구 데이터 읽기
src = openpyxl.load_workbook(SRC, data_only=True)
sws = src['아파트단지']
source_rows = []
for r in range(2, sws.max_row + 1):
    src_row = {
        '동': sws.cell(r, 3).value,
        '단지코드': sws.cell(r, 5).value,
        '단지명': sws.cell(r, 6).value,
        '사용승인일': sws.cell(r, 7).value,
        '총세대수': sws.cell(r, 8).value,
        '전용면적': sws.cell(r, 9).value,
        '구조': sws.cell(r, 10).value,
        '방수': sws.cell(r, 11).value,
        '욕실수': sws.cell(r, 12).value,
        '공급': sws.cell(r, 13).value,
        '평형별세대수': sws.cell(r, 14).value,
    }
    if src_row['단지코드'] is None:
        continue
    source_rows.append(src_row)
print(f'남구 파일 데이터 행: {len(source_rows)}')

# 대상 파일 열기
wb = openpyxl.load_workbook(DST)
ws = wb['부산 남구']

# 기존 5행부터 끝까지 내용 지우기 (템플릿 1행만 있었음)
for r in range(5, ws.max_row + 1):
    for c in range(1, 42):
        cell = ws.cell(r, c)
        cell.value = None
        cell.hyperlink = None

# 링크 폰트
link_font_base = None

# 데이터 기록
for i, d in enumerate(source_rows):
    r = 5 + i
    ws.cell(r, 1, value=d['동'])                    # A 동
    ws.cell(r, 4, value=d['단지코드'])                # D 단지코드
    name_cell = ws.cell(r, 5, value=d['단지명'])      # E 단지명
    # 하이퍼링크
    try:
        code_int = int(d['단지코드'])
        name_cell.hyperlink = URL.format(code_int)
        base = name_cell.font
        name_cell.font = Font(
            name=base.name, size=base.size, bold=base.bold,
            color='0563C1', underline='single',
        )
    except Exception:
        pass

    ws.cell(r, 6, value=d['사용승인일'])              # F 사용승인일
    ws.cell(r, 7, value=d['총세대수'])                # G 총세대수
    ws.cell(r, 8, value=f'=2026-F{r}+1')             # H 나이 수식

    # 공급면적 J
    try:
        j_val = int(float(d['공급']))
        ws.cell(r, 10, value=j_val)
        ws.cell(r, 12, value=round(j_val / PYONG_DIVISOR))  # L 평형
    except Exception:
        pass

    # 전용면적 K
    try:
        k_val = int(float(d['전용면적']))
        ws.cell(r, 11, value=k_val)
    except Exception:
        ws.cell(r, 11, value=d['전용면적'])

    ws.cell(r, 13, value=d['구조'])                   # M
    ws.cell(r, 14, value=d['방수'])                   # N
    ws.cell(r, 15, value=d['욕실수'])                  # O
    ws.cell(r, 16, value=d['평형별세대수'])            # P

wb.save(DST)
print(f'저장: {DST}')
print(f'부산 남구 탭에 {len(source_rows)}행 입력 완료')
