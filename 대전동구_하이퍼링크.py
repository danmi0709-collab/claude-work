"""대전동구 탭 단지명에 네이버 부동산 하이퍼링크 추가 (동구 (1).xlsx 참조)"""
import sys, io, os, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from openpyxl.styles import Font

BASE = r'C:\Users\한나\OneDrive\강성업무용\바탕 화면\문서\CLAUDE'
FILE = os.path.join(BASE, '시세시트_링크추가.xlsx')
REF = os.path.join(BASE, '동구 (1).xlsx')

URL = 'https://new.land.naver.com/complexes/{}?ms=2AfAQG,3zprE6,17&a=APT:ABYG:JGC&b=A1:B1&e=RETAIL&h=34&i=132&l=198&ad=true'


def norm(s):
    if not s: return ''
    return re.sub(r'[()\[\]{}·.\-_,/&+\s]', '', str(s).lower())


# 참조 파일에서 단지명 → 단지코드 매핑
print('동구 (1).xlsx 매핑 구축...')
wb_ref = openpyxl.load_workbook(REF, data_only=True)
ws_ref = wb_ref.worksheets[0]

# (읍면동, 정규화단지명) → 단지코드
code_map = {}
# 정규화단지명만으로도 매핑 (폴백)
name_only = {}

for r in range(2, ws_ref.max_row + 1):
    dong = ws_ref.cell(r, 3).value  # 읍면동
    code = ws_ref.cell(r, 5).value  # 단지코드
    name = ws_ref.cell(r, 6).value  # 단지명
    if not (code and name): continue
    try:
        code_int = int(code)
    except:
        continue
    nn = norm(name)
    dn = norm(dong) if dong else ''
    code_map[(dn, nn)] = code_int
    name_only[nn] = code_int

print(f'  매핑 건수: {len(code_map)} (고유 단지명: {len(name_only)})')


# 시세시트 대전동구 탭에 하이퍼링크 추가
wb = openpyxl.load_workbook(FILE)
ws = wb['대전동구']

added = not_found = 0
missing = []
for r in range(5, ws.max_row + 1):
    dong = ws.cell(r, 1).value  # A열
    name = ws.cell(r, 4).value  # D열 단지명
    if not name: continue
    nn = norm(name)
    dn = norm(dong) if dong else ''

    code = code_map.get((dn, nn)) or name_only.get(nn)
    if not code:
        not_found += 1
        if len(missing) < 10:
            missing.append((r, dong, name))
        continue

    cell = ws.cell(r, 4)
    cell.hyperlink = URL.format(code)
    ef = cell.font
    cell.font = Font(name=ef.name, size=ef.size, bold=ef.bold,
                     color='0563C1', underline='single')
    added += 1

wb.save(FILE)
print(f'\n하이퍼링크 추가: {added}개 / 미매칭: {not_found}개')
if missing:
    print('미매칭 샘플:')
    for r, d, n in missing:
        print(f'  행{r} [{d}] {n}')
