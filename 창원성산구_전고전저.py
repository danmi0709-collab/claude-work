"""창원성산구 탭 전고(U)/전저(V) 채우기
- 전고: 2021.01 ~ 2022.12 최고 매매가
- 전저: 2022.01 ~ 2023.12 최저 매매가
"""
import sys, io, os, re, csv, datetime
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from collections import defaultdict, Counter

BASE = r'C:\Users\한나\OneDrive\강성업무용\바탕 화면\문서\CLAUDE'
FILE = os.path.join(BASE, '시세시트_링크추가.xlsx')
CSV  = os.path.join(BASE, '창원성산구_시세', '매매_원본.csv')

HIGH_FROM = datetime.date(2021, 1, 1);  HIGH_TO = datetime.date(2022, 12, 31)
LOW_FROM  = datetime.date(2022, 1, 1);  LOW_TO  = datetime.date(2023, 12, 31)

def norm_name(s):
    if not s: return ''
    return re.sub(r'[()\[\]{}·.\-_,/&+]', '', re.sub(r'\s+', '', str(s).lower()))

def norm_dong(s):
    if not s: return ''
    s = str(s).strip()
    m = re.search(r'^(.+?[읍면동])(\s|$)', s)
    return m.group(1) if m else s

def parse_amt(s):
    if s is None: return None
    s = str(s).replace(',', '').strip()
    try: return int(s)
    except:
        try: return int(float(s))
        except: return None

# CSV 로드
trades = []
with open(CSV, 'r', encoding='utf-8-sig') as f:
    for r in csv.DictReader(f):
        try:
            y = int(r['dealYear']); m = int(r['dealMonth']); d = int(r['dealDay'])
            dt = datetime.date(y, m, d)
            area = float(r['excluUseAr'])
            by = int(r['buildYear']) if r['buildYear'] else None
        except: continue
        amt = parse_amt(r.get('dealAmount'))
        if amt is None or by is None: continue
        trades.append({
            'date': dt, 'dong': norm_dong(r.get('umdNm')),
            'apt': (r.get('aptNm') or '').strip(),
            'apt_norm': norm_name(r.get('aptNm')),
            'build_year': by, 'area_round': round(area), 'amount': amt,
        })
print(f'매매 {len(trades)}건')

idx = defaultdict(list)
for it in trades:
    idx[(it['dong'], it['build_year'], it['area_round'])].append(it)

def best_match(dong, byear, area_round, name, tol=1):
    cand = []
    for da in range(-tol, tol+1):
        cand.extend(idx.get((dong, byear, area_round + da), []))
    if not cand: return []
    wn = norm_name(name)
    exact = [c for c in cand if c['apt_norm'] == wn]
    if exact: return exact
    sub = [c for c in cand if wn and (wn in c['apt_norm'] or c['apt_norm'] in wn)]
    if sub:
        top = Counter(c['apt'] for c in sub).most_common(1)[0][0]
        return [c for c in sub if c['apt'] == top]
    top = Counter(c['apt'] for c in cand).most_common(1)[0][0]
    return [c for c in cand if c['apt'] == top]

wb = openpyxl.load_workbook(FILE)
ws = wb['창원성산구']

# 컬럼: A=동(1), E=단지명(5), F=사용승인일(6), K=전용면적(11), U=전고(21), V=전저(22)
high_n = low_n = proc = 0
for r in range(5, ws.max_row + 1):
    dong = ws.cell(r, 1).value
    name = ws.cell(r, 5).value
    y    = ws.cell(r, 6).value
    a    = ws.cell(r, 11).value
    if not (dong and name and y and a): continue
    try:
        byear = int(float(str(y)[:4]))
        ar = round(float(a))
    except: continue
    proc += 1
    items = best_match(norm_dong(str(dong)), byear, ar, str(name))
    if not items: continue
    high = max((i['amount'] for i in items if HIGH_FROM <= i['date'] <= HIGH_TO), default=None)
    low  = min((i['amount'] for i in items if LOW_FROM  <= i['date'] <= LOW_TO),  default=None)
    if high is not None:
        ws.cell(r, 21, value=round(high / 10000, 2))  # 억 단위
        high_n += 1
    if low is not None:
        ws.cell(r, 22, value=round(low / 10000, 2))   # 억 단위
        low_n += 1

wb.save(FILE)
print(f'처리 {proc}행 | 전고 {high_n}행 | 전저 {low_n}행')
