"""서울강서구 탭 생성 + 전고전저 채우기"""
import sys, io, os, re, csv, datetime
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, PatternFill
from collections import defaultdict, Counter
from copy import copy

BASE = r'C:\Users\한나\OneDrive\강성업무용\바탕 화면\문서\CLAUDE'
FILE = os.path.join(BASE, '시세시트_링크추가.xlsx')
SRC  = os.path.join(BASE, '강서구.xlsx')
CSV  = os.path.join(BASE, '강서구_시세', '매매_원본.csv')

URL   = 'https://new.land.naver.com/complexes/{}?ms=2AfAQG,3zprE6,17&a=APT:ABYG:JGC&b=A1:B1&e=RETAIL&h=34&i=132&l=198&ad=true'
PYONG = 3.3058
GRAY  = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
WHITE = PatternFill(fill_type=None)

HIGH_FROM = datetime.date(2021, 1, 1);  HIGH_TO = datetime.date(2022, 12, 31)
LOW_FROM  = datetime.date(2022, 1, 1);  LOW_TO  = datetime.date(2023, 12, 31)

def extract_num(s):
    if s is None: return None
    m = re.search(r'(\d+)', str(s))
    return int(m.group(1)) if m else None

# ── 1. 강서구.xlsx 읽기 ──
print('강서구.xlsx 로드...')
src_wb = openpyxl.load_workbook(SRC, data_only=True)
src_ws = src_wb.worksheets[0]
rows = []
for r in range(2, src_ws.max_row + 1):
    code = src_ws.cell(r, 5).value
    if code is None: continue
    try: code = int(code)
    except: continue
    rows.append({
        'dong':    src_ws.cell(r, 3).value,
        'code':    code,
        'name':    src_ws.cell(r, 6).value,
        'approved':src_ws.cell(r, 7).value,
        'hh_tot':  src_ws.cell(r, 8).value,
        'area_excl':src_ws.cell(r, 9).value,
        'struct':  src_ws.cell(r, 10).value,
        'rooms':   src_ws.cell(r, 11).value,
        'baths':   src_ws.cell(r, 12).value,
        'type_nm': src_ws.cell(r, 13).value,
        'hh_type': src_ws.cell(r, 14).value,
    })
print(f'  {len(rows)}행')

# ── 2. 서울강서구 탭 생성 ──
print('서울강서구 탭 생성...')
wb = openpyxl.load_workbook(FILE)
if '서울강서구' in wb.sheetnames:
    del wb['서울강서구']
tmpl = wb['창원성산구']
ws = wb.copy_worksheet(tmpl)
ws.title = '서울강서구'

for r in range(5, ws.max_row + 1):
    for c in range(1, 42):
        ws.cell(r, c).value = None
        ws.cell(r, c).hyperlink = None
        ws.cell(r, c).fill = WHITE

for i, d in enumerate(rows):
    r = 5 + i
    ws.cell(r, 1, value=d['dong'])
    ws.cell(r, 4, value=d['code'])
    nc = ws.cell(r, 5, value=d['name'])
    nc.hyperlink = URL.format(d['code'])
    nc.font = Font(name=nc.font.name, size=nc.font.size, bold=nc.font.bold,
                   color='0563C1', underline='single')
    ws.cell(r, 6, value=d['approved'])
    ws.cell(r, 7, value=d['hh_tot'])
    ws.cell(r, 8, value=f'=2026-INT(F{r})+1')
    supply = extract_num(d['type_nm'])
    if supply:
        ws.cell(r, 10, value=supply)
        ws.cell(r, 12, value=round(supply / PYONG))
    try: ws.cell(r, 11, value=int(float(d['area_excl'])))
    except: ws.cell(r, 11, value=d['area_excl'])
    ws.cell(r, 13, value=d['struct'])
    ws.cell(r, 14, value=d['rooms'])
    ws.cell(r, 15, value=d['baths'])
    ws.cell(r, 16, value=d['hh_type'])

# 배경색 토글
prev_key = None; use_gray = False
for i, d in enumerate(rows):
    r = 5 + i
    key = (d['dong'], d['name'], d['approved'])
    if key != prev_key:
        use_gray = not use_gray; prev_key = key
    if use_gray:
        for c in range(1, 38): ws.cell(r, c).fill = GRAY

wb.save(FILE)
print(f'  {len(rows)}행 입력 + 배경색 완료')

# ── 3. 전고전저 ──
print('전고전저 채우기...')

def norm_name(s):
    return re.sub(r'[()\[\]{}·.\-_,/&+]', '', re.sub(r'\s+', '', str(s).lower())) if s else ''
def norm_dong(s):
    if not s: return ''
    s = str(s).strip()
    m = re.search(r'^(.+?[읍면동])(\s|$)', s)
    return m.group(1) if m else s
def parse_amt(s):
    if s is None: return None
    s = str(s).replace(',', '').strip()
    try: return int(s)
    except:
        try: return int(float(s))
        except: return None

trades = []
with open(CSV, 'r', encoding='utf-8-sig') as f:
    for row in csv.DictReader(f):
        try:
            y=int(row['dealYear']); m=int(row['dealMonth']); d=int(row['dealDay'])
            dt=datetime.date(y,m,d); area=float(row['excluUseAr'])
            by=int(row['buildYear']) if row['buildYear'] else None
        except: continue
        amt=parse_amt(row.get('dealAmount'))
        if amt is None or by is None: continue
        trades.append({'date':dt,'dong':norm_dong(row.get('umdNm')),
            'apt':(row.get('aptNm') or '').strip(),
            'apt_norm':norm_name(row.get('aptNm')),
            'build_year':by,'area_round':round(area),'amount':amt})
print(f'  매매 {len(trades)}건')

idx = defaultdict(list)
for it in trades: idx[(it['dong'],it['build_year'],it['area_round'])].append(it)

def best_match(dong, byear, ar, name, tol=1):
    cand=[]
    for da in range(-tol,tol+1): cand.extend(idx.get((dong,byear,ar+da),[]))
    if not cand: return []
    wn=norm_name(name)
    exact=[c for c in cand if c['apt_norm']==wn]
    if exact: return exact
    sub=[c for c in cand if wn and (wn in c['apt_norm'] or c['apt_norm'] in wn)]
    if sub:
        top=Counter(c['apt'] for c in sub).most_common(1)[0][0]
        return [c for c in sub if c['apt']==top]
    top=Counter(c['apt'] for c in cand).most_common(1)[0][0]
    return [c for c in cand if c['apt']==top]

wb = openpyxl.load_workbook(FILE)
ws = wb['서울강서구']
high_n=low_n=proc=0
for r in range(5, ws.max_row+1):
    dong=ws.cell(r,1).value; name=ws.cell(r,5).value
    y_v=ws.cell(r,6).value; a_v=ws.cell(r,11).value
    if not (dong and name and y_v and a_v): continue
    try: byear=int(float(str(y_v)[:4])); ar=round(float(a_v))
    except: continue
    proc+=1
    items=best_match(norm_dong(str(dong)),byear,ar,str(name))
    if not items: continue
    high=max((i['amount'] for i in items if HIGH_FROM<=i['date']<=HIGH_TO),default=None)
    low =min((i['amount'] for i in items if LOW_FROM <=i['date']<=LOW_TO), default=None)
    if high is not None: ws.cell(r,21,value=round(high/10000,2)); high_n+=1
    if low  is not None: ws.cell(r,22,value=round(low /10000,2)); low_n+=1

wb.save(FILE)
print(f'  처리 {proc}행 | 전고 {high_n}행 | 전저 {low_n}행')
print('완료!')
