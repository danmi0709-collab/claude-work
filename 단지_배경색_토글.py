"""
천안동남구 시트: 행24부터 끝까지(255행) 단지가 바뀔 때마다 배경색 토글.
- 시작: 행23 목천신도브래뉴2차가 회색이었으므로, 다음 단지(행24 목천프라임)는 흰색.
- 이후 단지코드 바뀔 때마다 회색↔흰색 교대.
- 회색 스펙: 행9 휴먼시아와 동일 (theme=0, tint=-0.15), 칠할 열도 행9 그대로 복제.
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import Color
from copy import copy

FILE = r'C:\Users\한나\OneDrive\강성업무용\바탕 화면\문서\CLAUDE\시세시트_링크추가.xlsx'

wb = openpyxl.load_workbook(FILE)
ws = wb['천안동남구']

# 행9(휴먼시아 = 회색 단지) 에서 회색 칠한 열 목록과 실제 fill 객체 복제
REF_ROW = 9
gray_cols_fills = {}
for c in range(1, 52):
    cell = ws.cell(REF_ROW, c)
    if cell.fill.patternType == 'solid':
        gray_cols_fills[c] = copy(cell.fill)

print(f'회색 적용 열 수: {len(gray_cols_fills)}')

START_ROW = 24
END_ROW = 255

# 색 상태 추적: 이전 단지(행23)가 회색이었음
prev_code = ws.cell(23, 4).value
prev_was_gray = True  # 행23 회색

applied_rows = 0
for r in range(START_ROW, END_ROW + 1):
    code = ws.cell(r, 4).value
    if code is None:
        continue
    # 단지코드가 바뀔 때만 색 토글
    if code != prev_code:
        prev_was_gray = not prev_was_gray
        prev_code = code
    # 현재 단지가 회색이면 행9 패턴대로 칠하기
    if prev_was_gray:
        for col, fill in gray_cols_fills.items():
            ws.cell(r, col).fill = copy(fill)
        applied_rows += 1

wb.save(FILE)
print(f'회색 적용 행 수: {applied_rows}')
print(f'저장: {FILE}')
