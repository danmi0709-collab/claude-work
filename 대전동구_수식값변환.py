"""대전동구 탭의 빈칸채우기 수식(=B5 등)을 실제 값으로 변환"""
import sys, io, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

FILE = r'C:\Users\한나\OneDrive\강성업무용\바탕 화면\문서\CLAUDE\시세시트_링크추가.xlsx'

# data_only=True로 수식 계산값 읽기
wb_val = openpyxl.load_workbook(FILE, data_only=True)
ws_val = wb_val['대전동구']

# 일반 로드로 수식 셀 수정
wb = openpyxl.load_workbook(FILE)
ws = wb['대전동구']

# 단순 셀 참조 수식 패턴: =A1, =BC123 등
simple_ref = re.compile(r'^=([A-Z]+)(\d+)$')

converted = 0
for row in ws.iter_rows():
    for cell in row:
        v = cell.value
        if v and isinstance(v, str) and v.startswith('='):
            m = simple_ref.match(v.strip())
            if m:
                # data_only 워크북에서 실제 값 가져오기
                actual = ws_val.cell(cell.row, cell.column).value
                if actual is not None:
                    cell.value = actual
                    converted += 1

wb.save(FILE)
print(f'변환 완료: {converted}개 수식 → 실제 값으로 교체')
