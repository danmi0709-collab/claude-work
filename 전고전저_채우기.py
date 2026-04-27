"""
시세시트의 천안동남구 시트에서 각 행에 전고/전저 채우기.
- 전고: 2021.01.01 ~ 2022.12.31 매매 최고가 → U열
- 전저: 2022.01.01 ~ 2023.12.31 매매 최저가 → V열
매칭키: (동, 건축년도, 반올림 전용면적) + 단지명 정규화 유사도
실거래 데이터: 천안동남구_시세/매매_원본.csv + 매매_추가.csv
"""
import sys, io, os, re, csv, datetime
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from collections import defaultdict, Counter

BASE = r'C:\Users\한나\OneDrive\강성업무용\바탕 화면\문서\CLAUDE'
FILE = os.path.join(BASE, '시세시트_링크추가.xlsx')
DATA = os.path.join(BASE, '천안동남구_시세')

HIGH_FROM = datetime.date(2021, 1, 1)
HIGH_TO   = datetime.date(2022, 12, 31)
LOW_FROM  = datetime.date(2022, 1, 1)
LOW_TO    = datetime.date(2023, 12, 31)


def norm_name(s):
    if not s: return ''
    s = str(s).lower()
    s = re.sub(r'\s+', '', s)
    s = re.sub(r'[()\[\]{}·.\-_,/&+]', '', s)
    return s


def norm_dong(s):
    if not s: return ''
    s = str(s).strip()
    m = re.search(r'^(.+?[읍면동])(\s|$)', s)
    return m.group(1) if m else s


def parse_amt(s):
    if s is None: return None
    s = str(s).replace(',', '').strip()
    if not s: return None
    try: return int(s)
    except:
        try: return int(float(s))
        except: return None


def load(path):
    with open(path, 'r', encoding='utf-8-sig') as f:
        return list(csv.DictReader(f))


print('실거래 CSV 로드...')
rows = load(os.path.join(DATA, '매매_원본.csv')) + load(os.path.join(DATA, '매매_추가.csv'))
trades = []
for r in rows:
    try:
        y = int(r['dealYear']); m = int(r['dealMonth']); d = int(r['dealDay'])
        dt = datetime.date(y, m, d)
        area = float(r['excluUseAr'])
        by = int(r['buildYear']) if r['buildYear'] else None
    except: continue
    amt = parse_amt(r.get('dealAmount'))
    if amt is None or by is None: continue
    trades.append({
        'date': dt, 'dong': norm_dong(r.get('umdNm')),
        'apt': (r.get('aptNm') or '').strip(),
        'apt_norm': norm_name(r.get('aptNm')),
        'build_year': by, 'area': area, 'area_round': round(area), 'amount': amt,
    })
print(f'  매매 {len(trades)}건')

idx = defaultdict(list)
for it in trades:
    idx[(it['dong'], it['build_year'], it['area_round'])].append(it)


def best_match(dong, byear, area_round, wolbu_name, tol=1):
    cand = []
    for da in range(-tol, tol+1):
        cand.extend(idx.get((dong, byear, area_round + da), []))
    if not cand: return [], 'none'
    wn = norm_name(wolbu_name)
    exact = [c for c in cand if c['apt_norm'] == wn]
    if exact: return exact, 'exact'
    sub = [c for c in cand if wn and (wn in c['apt_norm'] or c['apt_norm'] in wn)]
    if sub:
        top = Counter(c['apt'] for c in sub).most_common(1)[0][0]
        return [c for c in sub if c['apt'] == top], 'substring'
    top = Counter(c['apt'] for c in cand).most_common(1)[0][0]
    return [c for c in cand if c['apt'] == top], 'fallback'


wb = openpyxl.load_workbook(FILE)
ws = wb['천안동남구']

COL_DONG=1; COL_NAME=5; COL_YEAR=6; COL_AREA=11; COL_HIGH=21; COL_LOW=22

high_filled = low_filled = rows_processed = 0
none_rows = []
for r in range(5, ws.max_row+1):
    dong = ws.cell(r, COL_DONG).value
    name = ws.cell(r, COL_NAME).value
    year_val = ws.cell(r, COL_YEAR).value
    area_val = ws.cell(r, COL_AREA).value
    if not (dong and name and year_val and area_val): continue
    try:
        byear = int(float(year_val))  # 2006.05 → 2006
        area_round = round(float(area_val))
    except: continue
    rows_processed += 1

    items, flag = best_match(norm_dong(dong), byear, area_round, str(name))
    if flag == 'none':
        none_rows.append((r, dong, name, byear, area_val))
        continue
    high = max((i['amount'] for i in items if HIGH_FROM <= i['date'] <= HIGH_TO), default=None)
    low = min((i['amount'] for i in items if LOW_FROM <= i['date'] <= LOW_TO), default=None)
    if high is not None:
        ws.cell(r, COL_HIGH, value=high); high_filled += 1
    if low is not None:
        ws.cell(r, COL_LOW, value=low); low_filled += 1

wb.save(FILE)
print(f'처리 행: {rows_processed}')
print(f'전고 채움: {high_filled}')
print(f'전저 채움: {low_filled}')
print(f'매칭 실패: {len(none_rows)}')
if none_rows[:5]:
    print('실패 예시:')
    for r, d, n, y, a in none_rows[:10]:
        print(f'  행{r} {d} {n} {y} 전용{a}')
print(f'저장: {FILE}')
