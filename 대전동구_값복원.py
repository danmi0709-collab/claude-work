"""대전동구 탭을 백업에서 '값만' 복원 (수식 대신 계산된 값 사용)"""
import sys, io, os, re, csv
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from openpyxl.cell import MergedCell
from openpyxl.styles import PatternFill
from collections import defaultdict, Counter
from statistics import mean
import copy

BASE = r'C:\Users\한나\OneDrive\강성업무용\바탕 화면\문서\CLAUDE'
FILE = os.path.join(BASE, '시세시트_링크추가.xlsx')
BACKUP = os.path.join(BASE, '시세시트_원본백업.xlsx')
DATA = os.path.join(BASE, '대전동구_시세')

# ── 1단계: 백업에서 값만 복원 (data_only=True 사용) ──
print('1단계: 백업에서 값 추출 중...')
wb_bak_val = openpyxl.load_workbook(BACKUP, data_only=True)
wb_bak_style = openpyxl.load_workbook(BACKUP)  # 스타일용

ws_bak_val = wb_bak_val['대전동구']
ws_bak_style = wb_bak_style['대전동구']

max_r = ws_bak_val.max_row
max_c = ws_bak_val.max_column
print(f'  백업 범위: {max_r}행 x {max_c}열')

# 현재 파일 로드
wb = openpyxl.load_workbook(FILE)
ws = wb['대전동구']

# 기존 병합셀 해제 (오류 시 건너뜀)
merge_ranges = list(ws.merged_cells.ranges)
for mr in merge_ranges:
    try:
        ws.unmerge_cells(str(mr))
    except Exception:
        pass

# 기존 내용 초기화 (값과 스타일)
for r in range(1, ws.max_row + 1):
    for c in range(1, ws.max_column + 1):
        ws.cell(r, c).value = None
        ws.cell(r, c).hyperlink = None

# 백업에서 값 + 스타일 복사
for r in range(1, max_r + 1):
    for c in range(1, max_c + 1):
        src_val = ws_bak_val.cell(r, c)
        src_style = ws_bak_style.cell(r, c)
        dst = ws.cell(r, c)

        if isinstance(src_val, MergedCell):
            continue

        # 값: data_only 버전에서 (수식 대신 계산값)
        dst.value = src_val.value

        # 스타일: 일반 버전에서
        if src_style.has_style:
            dst.font = copy.copy(src_style.font)
            dst.fill = copy.copy(src_style.fill)
            dst.border = copy.copy(src_style.border)
            dst.alignment = copy.copy(src_style.alignment)
            dst.number_format = src_style.number_format

        if src_style.hyperlink:
            dst.hyperlink = copy.copy(src_style.hyperlink)

# 병합셀 복원 (헤더 영역만 - 데이터 영역은 병합 안 함)
for mr in ws_bak_style.merged_cells.ranges:
    # 4행 이하의 헤더 영역만 병합 유지
    if mr.max_row <= 4:
        try:
            ws.merge_cells(str(mr))
        except Exception as e:
            print(f'  병합 실패 {mr}: {e}')

# 행 높이, 열 너비
for r in range(1, max_r + 1):
    if ws_bak_style.row_dimensions[r].height:
        ws.row_dimensions[r].height = ws_bak_style.row_dimensions[r].height

from openpyxl.utils import get_column_letter
for c in range(1, max_c + 1):
    cl = get_column_letter(c)
    if ws_bak_style.column_dimensions[cl].width:
        ws.column_dimensions[cl].width = ws_bak_style.column_dimensions[cl].width

wb.save(FILE)
print('  복원 완료')


# ── 2단계: 단지 토글 배경색 적용 ──
print('\n2단계: 단지 토글 배경색 적용...')
wb = openpyxl.load_workbook(FILE)
ws = wb['대전동구']

GRAY = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
WHITE = PatternFill(fill_type=None)

prev_key = None
use_gray = False
for r in range(5, ws.max_row + 1):
    dong = ws.cell(r, 1).value
    name = ws.cell(r, 4).value
    year = ws.cell(r, 8).value
    if not (dong and name and year): continue
    key = (dong, name, year)
    if key != prev_key:
        use_gray = not use_gray
        prev_key = key
    if use_gray:
        for c in range(1, 36):
            if not isinstance(ws.cell(r, c), MergedCell):
                ws.cell(r, c).fill = GRAY

wb.save(FILE)
print('  배경색 적용 완료')


# ── 3단계: X/Y/AA/AB 채우기 ──
print('\n3단계: X/Y/AA/AB 채우기...')

RATIO_THRESHOLD = 0.80
ROLL = 6

def norm_name(s):
    if not s: return ''
    return re.sub(r'[()\[\]{}·.\-_,/&+]', '', re.sub(r'\s+', '', str(s).lower()))

def norm_dong(s):
    if not s: return ''
    s = str(s).strip()
    m = re.search(r'^(.+?[읍면동])(\s|$)', s)
    return m.group(1) if m else s

def parse_amt(s):
    if s is None: return None
    s = str(s).replace(',', '').strip()
    if not s: return None
    try: return int(s)
    except:
        try: return int(float(s))
        except: return None

def load_csv(p):
    with open(p, 'r', encoding='utf-8-sig') as f:
        return list(csv.DictReader(f))

def prepare(rows, amt_key, is_rent=False):
    out = []
    for r in rows:
        try:
            y = int(r['dealYear']); m = int(r['dealMonth'])
            area = float(r['excluUseAr'])
            by = int(r['buildYear']) if r['buildYear'] else None
        except: continue
        amt = parse_amt(r.get(amt_key))
        if amt is None or by is None: continue
        if is_rent and (r.get('contractType') or '').strip() == '갱신': continue
        out.append({
            'ym': (y, m), 'dong': norm_dong(r.get('umdNm')),
            'apt': (r.get('aptNm') or '').strip(),
            'apt_norm': norm_name(r.get('aptNm')),
            'build_year': by, 'area_round': round(area), 'amount': amt,
        })
    return out

trades = prepare(load_csv(os.path.join(DATA, '매매_원본.csv')), 'dealAmount')
rents  = prepare(load_csv(os.path.join(DATA, '전세_원본.csv')), 'deposit', is_rent=True)

def build_idx(items):
    idx = defaultdict(list)
    for it in items:
        idx[(it['dong'], it['build_year'], it['area_round'])].append(it)
    return idx

trade_idx = build_idx(trades)
rent_idx = build_idx(rents)

def best_match(idx, dong, byear, area_round, name, tol=1):
    cand = []
    for da in range(-tol, tol+1):
        cand.extend(idx.get((dong, byear, area_round + da), []))
    if not cand: return []
    wn = norm_name(name)
    exact = [c for c in cand if c['apt_norm'] == wn]
    if exact: return exact
    sub = [c for c in cand if wn and (wn in c['apt_norm'] or c['apt_norm'] in wn)]
    if sub:
        top = Counter(c['apt'] for c in sub).most_common(1)[0][0]
        return [c for c in sub if c['apt'] == top]
    top = Counter(c['apt'] for c in cand).most_common(1)[0][0]
    return [c for c in cand if c['apt'] == top]

def monthly_items(items):
    b = defaultdict(list)
    for it in items: b[it['ym']].append(it['amount'])
    return b

def ym_next(y, m): return (y, m+1) if m < 12 else (y+1, 1)
def ym_prev(y, m): return (y, m-1) if m > 1 else (y-1, 12)
def ym_shift(ym, d):
    y, m = ym
    for _ in range(abs(d)):
        if d > 0: y, m = ym_next(y, m)
        else: y, m = ym_prev(y, m)
    return (y, m)

def rolling_avg(bucket, months):
    out = {}
    for ym in months:
        vals = []
        for k in range(ROLL):
            vals.extend(bucket.get(ym_shift(ym, -k), []))
        if vals: out[ym] = mean(vals)
    return out

def find_best_period(t, r):
    tb = monthly_items(t); rb = monthly_items(r)
    ay = sorted(set(tb) | set(rb))
    if not ay: return None
    months = []; cur = ay[0]
    while cur <= ay[-1]:
        months.append(cur); cur = ym_next(*cur)
    ta = rolling_avg(tb, months); ra = rolling_avg(rb, months)
    common = [ym for ym in months if ym in ta and ym in ra and ta[ym] > 0]
    if not common: return None
    ratios = {ym: ra[ym]/ta[ym] for ym in common}
    high = [ym for ym in sorted(ratios) if ratios[ym] >= RATIO_THRESHOLD]
    if not high: return None
    runs = []; cur_run = [high[0]]
    for ym in high[1:]:
        if ym_next(*cur_run[-1]) == ym: cur_run.append(ym)
        else: runs.append(cur_run); cur_run = [ym]
    runs.append(cur_run)
    best = None
    for run in runs:
        mn = min(run, key=lambda y: ta[y])
        if best is None or ta[mn] < best[0]: best = (ta[mn], run, mn)
    mt, run, mn = best
    return (mn, round(mt), round(ra[mn]), len(run))

def months_to_display(n):
    y, x = divmod(n, 12)
    return y + x / 100.0

wb = openpyxl.load_workbook(FILE)
ws = wb['대전동구']

C_DONG, C_NAME, C_YEAR, C_AREA = 1, 4, 8, 12  # L=전용면적㎡
C_X, C_Y, C_AA, C_AB = 24, 25, 27, 28

proc = filled = 0
for r in range(5, ws.max_row + 1):
    dong = ws.cell(r, C_DONG).value
    name = ws.cell(r, C_NAME).value
    y_val = ws.cell(r, C_YEAR).value
    a_val = ws.cell(r, C_AREA).value
    if not (dong and name and y_val and a_val): continue
    try:
        byear = int(float(str(y_val)[:4])); ar = round(float(a_val))
    except: continue
    proc += 1
    ti = best_match(trade_idx, norm_dong(str(dong)), byear, ar, str(name))
    ri = best_match(rent_idx, norm_dong(str(dong)), byear, ar, str(name))
    if ti and ri:
        res = find_best_period(ti, ri)
        if res:
            min_ym, min_t, rent_at, length = res
            ws.cell(r, C_X, value=min_ym[0] + min_ym[1]/100.0)
            ws.cell(r, C_Y, value=months_to_display(length))
            ws.cell(r, C_AA, value=min_t)
            ws.cell(r, C_AB, value=rent_at)
            filled += 1

wb.save(FILE)
print(f'  처리 {proc}행, 채움 {filled}행')
print('\n완료!')
