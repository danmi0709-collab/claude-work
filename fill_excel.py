# -*- coding: utf-8 -*-
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from pathlib import Path
import openpyxl

path = Path(r'C:\Users\한나\OneDrive\월부\202605 실전 송파구\(송파구) 매임 체크리스트.xlsx')
wb = openpyxl.load_workbook(path)
ws = wb.active

# 컬럼 D=물건2, E=물건3, F=물건4, G=물건5, H=물건6, I=물건7, J=물건8
# (C=물건1: 249-1401 이미 입력됨)

properties = [
    # (동호수, 매매가, 아파트명, 특이사항, 전세가, 만기일, 거주자)
    ('148-1403', 33.5, '잠실리센츠', '앞뒤트인뷰 27만기 10.5억 가성비 한강X', 10.5, '2027년', '세입자'),
    ('148-2001', 35,   '잠실리센츠', '앞뒤트인 한강뷰 28.1만기 8억', 8,    '2028년1월', '세입자'),
    ('120-1702', 36,   '잠실리센츠', '초역세권 남향 입주매매', None, None, '공실'),
    ('236-1402', 35,   '리센츠',     '입주 한강뷰', None, None, '공실'),
    ('203-2401', 35.5, '리센츠',     '입매매 통창 한강뷰', None, None, '집주인'),
    ('201-2401', 36.5, '리센츠',     '통창 한강뷰 27.4만기 11억', 11,   '2027년4월', '세입자'),
    ('327-1704', 33.5, '트리지움',   '입주 최근수리 여자세입자 이사비', None, None, '세입자'),
]

# 시작 컬럼: D = 4
start_col = 4  # D열

for i, (dong_ho, price, apt_name, memo, jeonse, expire, resident) in enumerate(properties):
    col = start_col + i
    col_letter = ws.cell(row=1, column=col).column_letter

    ws.cell(row=3, column=col).value = apt_name      # 아파트명
    ws.cell(row=5, column=col).value = '잠실정원공인'  # 부동산
    ws.cell(row=11, column=col).value = dong_ho       # 동호수
    ws.cell(row=15, column=col).value = resident      # 거주자
    ws.cell(row=18, column=col).value = memo          # 특이사항
    ws.cell(row=19, column=col).value = price         # 매매가
    if jeonse:
        ws.cell(row=21, column=col).value = jeonse    # 전세가
    if expire:
        ws.cell(row=17, column=col).value = expire    # 만기일

    print(f"  {col_letter}열 ({apt_name} {dong_ho}): 매매가 {price}억 입력 완료")

# C열 아파트명도 추가
ws['C3'] = '리센츠'

wb.save(path)
print("\n저장 완료!")
