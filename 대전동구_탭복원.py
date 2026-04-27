"""백업 파일에서 대전동구 탭을 현재 파일로 복원 후 X/Y/AA/AB 재채우기"""
import sys, io, os, re, csv
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from collections import defaultdict, Counter
from statistics import mean
import copy

BASE = r'C:\Users\한나\OneDrive\강성업무용\바탕 화면\문서\CLAUDE'
FILE = os.path.join(BASE, '시세시트_링크추가.xlsx')
BACKUP = os.path.join(BASE, '시세시트_원본백업.xlsx')
DATA = os.path.join(BASE, '대전동구_시세')

# ── 1단계: 대전동구 탭 복원 ──
print('1단계: 백업에서 대전동구 탭 복원 중...')
wb_bak = openpyxl.load_workbook(BACKUP)
wb = openpyxl.load_workbook(FILE)

ws_bak = wb_bak['대전동구']
ws = wb['대전동구']

# 기존 탭 내용 전체 삭제 후 백업 내용 복사
# 행/열 범위 파악
max_r = ws_bak.max_row
max_c = ws_bak.max_column

from openpyxl.cell import MergedCell

# 병합 셀 정보 복사
for merge in ws_bak.merged_cells.ranges:
    ws.merge_cells(str(merge))

for r in range(1, max_r + 1):
    for c in range(1, max_c + 1):
        src = ws_bak.cell(r, c)
        dst = ws.cell(r, c)
        if isinstance(src, MergedCell):
            continue  # 병합된 종속 셀은 건너뜀
        dst.value = src.value
        if src.has_style:
            dst.font = copy.copy(src.font)
            dst.fill = copy.copy(src.fill)
            dst.border = copy.copy(src.border)
            dst.alignment = copy.copy(src.alignment)
            dst.number_format = src.number_format
        if src.hyperlink:
            dst.hyperlink = src.hyperlink

# 행 높이, 열 너비 복사
for r in range(1, max_r + 1):
    if ws_bak.row_dimensions[r].height:
        ws.row_dimensions[r].height = ws_bak.row_dimensions[r].height

for c in range(1, max_c + 1):
    col_letter = get_column_letter(c)
    if ws_bak.column_dimensions[col_letter].width:
        ws.column_dimensions[col_letter].width = ws_bak.column_dimensions[col_letter].width

wb.save(FILE)
print(f'  복원 완료: {max_r}행 x {max_c}열')


# ── 2단계: X/Y/AA/AB 재채우기 ──
print('\n2단계: X/Y/AA/AB 재채우기...')

RATIO_THRESHOLD = 0.80
ROLL = 6

def norm_name(s):
    if not s: return ''
    return re.sub(r'[()\[\]{}·.\-_,/&+]', '', re.sub(r'\s+', '', str(s).lower()))

def norm_dong(s):
    if not s: return ''
    s = str(s).strip()
    m = re.search(r'^(.+?[읍면동])(\s|$)', s)
    return m.group(1) if m else s

def parse_amt(s):
    if s is None: return None
    s = str(s).replace(',', '').strip()
    if not s: return None
    try: return int(s)
    except:
        try: return int(float(s))
        except: return None

def load_csv(p):
    with open(p, 'r', encoding='utf-8-sig') as f:
        return list(csv.DictReader(f))

def prepare(rows, amt_key, is_rent=False):
    out = []
    for r in rows:
        try:
            y = int(r['dealYear']); m = int(r['dealMonth'])
            area = float(r['excluUseAr'])
            by = int(r['buildYear']) if r['buildYear'] else None
        except: continue
        amt = parse_amt(r.get(amt_key))
        if amt is None or by is None: continue
        if is_rent and (r.get('contractType') or '').strip() == '갱신': continue
        out.append({
            'ym': (y, m), 'dong': norm_dong(r.get('umdNm')),
            'apt': (r.get('aptNm') or '').strip(),
            'apt_norm': norm_name(r.get('aptNm')),
            'build_year': by, 'area_round': round(area), 'amount': amt,
        })
    return out

trades = prepare(load_csv(os.path.join(DATA, '매매_원본.csv')), 'dealAmount')
rents  = prepare(load_csv(os.path.join(DATA, '전세_원본.csv')), 'deposit', is_rent=True)
print(f'  매매 {len(trades)}건 / 전세 {len(rents)}건')

def build_idx(items):
    idx = defaultdict(list)
    for it in items:
        idx[(it['dong'], it['build_year'], it['area_round'])].append(it)
    return idx

trade_idx = build_idx(trades)
rent_idx = build_idx(rents)

def best_match(idx, dong, byear, area_round, name, tol=1):
    cand = []
    for da in range(-tol, tol+1):
        cand.extend(idx.get((dong, byear, area_round + da), []))
    if not cand: return []
    wn = norm_name(name)
    exact = [c for c in cand if c['apt_norm'] == wn]
    if exact: return exact
    sub = [c for c in cand if wn and (wn in c['apt_norm'] or c['apt_norm'] in wn)]
    if sub:
        top = Counter(c['apt'] for c in sub).most_common(1)[0][0]
        return [c for c in sub if c['apt'] == top]
    top = Counter(c['apt'] for c in cand).most_common(1)[0][0]
    return [c for c in cand if c['apt'] == top]

def monthly_items(items):
    b = defaultdict(list)
    for it in items: b[it['ym']].append(it['amount'])
    return b

def ym_next(y, m): return (y, m+1) if m < 12 else (y+1, 1)
def ym_prev(y, m): return (y, m-1) if m > 1 else (y-1, 12)
def ym_shift(ym, d):
    y, m = ym
    for _ in range(abs(d)):
        if d > 0: y, m = ym_next(y, m)
        else: y, m = ym_prev(y, m)
    return (y, m)

def rolling_avg(bucket, months):
    out = {}
    for ym in months:
        vals = []
        for k in range(ROLL):
            vals.extend(bucket.get(ym_shift(ym, -k), []))
        if vals: out[ym] = mean(vals)
    return out

def find_best_period(t, r):
    tb = monthly_items(t); rb = monthly_items(r)
    ay = sorted(set(tb) | set(rb))
    if not ay: return None
    months = []; cur = ay[0]
    while cur <= ay[-1]:
        months.append(cur); cur = ym_next(*cur)
    ta = rolling_avg(tb, months); ra = rolling_avg(rb, months)
    common = [ym for ym in months if ym in ta and ym in ra and ta[ym] > 0]
    if not common: return None
    ratios = {ym: ra[ym]/ta[ym] for ym in common}
    high = [ym for ym in sorted(ratios) if ratios[ym] >= RATIO_THRESHOLD]
    if not high: return None
    runs = []; cur_run = [high[0]]
    for ym in high[1:]:
        if ym_next(*cur_run[-1]) == ym: cur_run.append(ym)
        else: runs.append(cur_run); cur_run = [ym]
    runs.append(cur_run)
    best = None
    for run in runs:
        mn = min(run, key=lambda y: ta[y])
        if best is None or ta[mn] < best[0]: best = (ta[mn], run, mn)
    mt, run, mn = best
    return (mn, round(mt), round(ra[mn]), len(run))

def months_to_display(n):
    y, x = divmod(n, 12)
    return y + x / 100.0

wb = openpyxl.load_workbook(FILE)
ws = wb['대전동구']

C_DONG, C_NAME, C_YEAR, C_AREA = 1, 4, 8, 12  # L열=전용면적(㎡)
C_X, C_Y, C_AA, C_AB = 24, 25, 27, 28

proc = filled = 0
for r in range(5, ws.max_row + 1):
    dong = ws.cell(r, C_DONG).value
    name = ws.cell(r, C_NAME).value
    y_val = ws.cell(r, C_YEAR).value
    a_val = ws.cell(r, C_AREA).value
    if not (dong and name and y_val and a_val): continue
    if str(name).startswith('='): continue
    try:
        byear = int(float(str(y_val)[:4])); ar = round(float(a_val))
    except: continue
    proc += 1
    ti = best_match(trade_idx, norm_dong(str(dong)), byear, ar, str(name))
    ri = best_match(rent_idx, norm_dong(str(dong)), byear, ar, str(name))
    if ti and ri:
        res = find_best_period(ti, ri)
        if res:
            min_ym, min_t, rent_at, length = res
            ws.cell(r, C_X, value=min_ym[0] + min_ym[1]/100.0)
            ws.cell(r, C_Y, value=months_to_display(length))
            ws.cell(r, C_AA, value=min_t)
            ws.cell(r, C_AB, value=rent_at)
            filled += 1

wb.save(FILE)
print(f'처리 {proc}행, X/Y/AA/AB 채움 {filled}행')
print('\n완료!')
