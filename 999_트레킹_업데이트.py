"""9.9.9. 탭 트레킹 단지 자동 업데이트
- 9.9.9. 탭 행4 수식(=지역탭!E{행})에서 탭명+행번호 추출
- 행5~8 비어있으면 자동 채우기 + 84㎡ 행 빨간 표시
- 사용: python 999_트레킹_업데이트.py [지역탭명]
  예)  python 999_트레킹_업데이트.py 성동구
  인자 없으면 전체 처리
"""
import sys, io, re, openpyxl
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl.styles import PatternFill

BASE = r'C:\Users\한나\OneDrive\강성업무용\바탕 화면\문서\CLAUDE'
FILE = BASE + r'\시세시트_링크추가.xlsx'
RED  = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

# 행 번호 상수
ROW_NAME  = 4   # 단지명 (수식 =탭!E{row})
ROW_HH    = 5   # 세대수
ROW_YEAR  = 6   # 연식
ROW_PYONG = 7   # 평형(84)
ROW_HIGH  = 8   # 전고점(84)

target = sys.argv[1] if len(sys.argv) > 1 else None

wb = openpyxl.load_workbook(FILE)
ws9 = wb['9.9.9.']

total_updated = 0
total_red = 0

for c in range(1, ws9.max_column + 1):
    row4_val = ws9.cell(ROW_NAME, c).value
    if row4_val is None:
        continue

    # 행5 이미 채워진 경우 건너뜀
    if ws9.cell(ROW_HH, c).value is not None:
        continue

    # 행4 수식 파싱: =탭명!E행번호 or ='탭명'!E행번호
    m = re.match(r"^='?([^!']+)'?!E(\d+)$", str(row4_val).strip())
    if not m:
        # 수식이 아니면 건너뜀 (단지명 텍스트 직접 입력 방식은 미지원)
        continue

    region_tab = m.group(1)
    base_row   = int(m.group(2))

    # 특정 지역 지정 시 해당 탭만 처리
    if target and region_tab != target:
        continue

    if region_tab not in wb.sheetnames:
        print(f'col{c}: [{region_tab}] 탭 없음 → 건너뜀')
        continue

    ws_r = wb[region_tab]
    unit_name = ws_r.cell(base_row, 5).value  # E열 = 단지명

    if not unit_name:
        print(f'col{c}: {region_tab} 행{base_row} 단지명 없음 → 건너뜀')
        continue

    # 84㎡ 행 찾기 — 탭 전체에서 같은 단지명 + 전용면적 83~85
    row_84 = None
    for r in range(5, ws_r.max_row + 1):
        if ws_r.cell(r, 5).value != unit_name:
            continue
        k = ws_r.cell(r, 11).value  # K열 = 전용면적
        if k:
            try:
                if 83 <= float(str(k)) <= 85:
                    row_84 = r
                    break
            except: pass

    if row_84 is None:
        print(f'col{c} [{unit_name}]: 84㎡ 행 없음 → 기준행({base_row})으로 대체')
        row_84 = base_row

    # 9.9.9. 수식 입력
    ws9.cell(ROW_HH,    c, value=f'={region_tab}!G{base_row}')
    ws9.cell(ROW_YEAR,  c, value=f'=INT({region_tab}!F{base_row})')
    ws9.cell(ROW_PYONG, c, value=f'={region_tab}!L{row_84}')
    ws9.cell(ROW_HIGH,  c, value=f'={region_tab}!U{row_84}*10000')

    # 지역탭 84㎡ 행 빨간 배경색
    for col_r in range(1, ws_r.max_column + 1):
        ws_r.cell(row_84, col_r).fill = RED

    print(f'col{c} [{region_tab}] [{unit_name}] 기준행={base_row} 84㎡행={row_84} ✓')
    total_updated += 1
    total_red += 1

wb.save(FILE)
print(f'\n완료: {total_updated}개 단지 수식 입력 | {total_red}개 행 빨간색 표시')
