"""
부산수영구 탭: 데이터 채우기 + 배경색 토글.
(전고/전저는 별도 스크립트에서 실거래가 수집 후 진행)
"""
import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from openpyxl.styles import Font
from copy import copy

BASE = r'C:\Users\한나\OneDrive\강성업무용\바탕 화면\문서\CLAUDE'
SRC = os.path.join(BASE, '수영구.xlsx')
DST = os.path.join(BASE, '시세시트_링크추가.xlsx')
URL = 'https://new.land.naver.com/complexes/{}?ms=2AfAQG,3zprE6,17&a=APT:ABYG:JGC&b=A1:B1&e=RETAIL&h=34&i=132&l=198&ad=true'
PD = 3.3058

# 수영구 원본 로드
src = openpyxl.load_workbook(SRC, data_only=True)
sws = src['아파트단지']
rows = []
for r in range(2, sws.max_row + 1):
    d = {
        '동': sws.cell(r, 3).value,
        '단지코드': sws.cell(r, 5).value,
        '단지명': sws.cell(r, 6).value,
        '사용승인일': sws.cell(r, 7).value,
        '총세대수': sws.cell(r, 8).value,
        '전용면적': sws.cell(r, 9).value,
        '구조': sws.cell(r, 10).value,
        '방수': sws.cell(r, 11).value,
        '욕실수': sws.cell(r, 12).value,
        '공급': sws.cell(r, 13).value,
        '평형별세대수': sws.cell(r, 14).value,
    }
    if d['단지코드'] is None: continue
    rows.append(d)
print(f'수영구 원본 {len(rows)}행')

wb = openpyxl.load_workbook(DST)
ws = wb['부산수영구']

# 기존 5행부터 내용 지우기
for r in range(5, ws.max_row + 1):
    for c in range(1, 42):
        cell = ws.cell(r, c)
        cell.value = None
        cell.hyperlink = None

# 데이터 기록
for i, d in enumerate(rows):
    r = 5 + i
    ws.cell(r, 1, value=d['동'])
    ws.cell(r, 4, value=d['단지코드'])
    name_cell = ws.cell(r, 5, value=d['단지명'])
    try:
        code_int = int(d['단지코드'])
        name_cell.hyperlink = URL.format(code_int)
        b = name_cell.font
        name_cell.font = Font(name=b.name, size=b.size, bold=b.bold,
                              color='0563C1', underline='single')
    except Exception: pass
    ws.cell(r, 6, value=d['사용승인일'])
    ws.cell(r, 7, value=d['총세대수'])
    ws.cell(r, 8, value=f'=2026-F{r}+1')
    try:
        j = int(float(d['공급']))
        ws.cell(r, 10, value=j)
        ws.cell(r, 12, value=round(j / PD))
    except Exception: pass
    try:
        k = int(float(d['전용면적']))
        ws.cell(r, 11, value=k)
    except Exception:
        ws.cell(r, 11, value=d['전용면적'])
    ws.cell(r, 13, value=d['구조'])
    ws.cell(r, 14, value=d['방수'])
    ws.cell(r, 15, value=d['욕실수'])
    ws.cell(r, 16, value=d['평형별세대수'])
print(f'데이터 {len(rows)}행 입력')

# 배경색 토글 (천안동남구 행9 회색 패턴 복제)
ref_ws = wb['천안동남구']
gray_fills = {}
for c in range(1, 52):
    cell = ref_ws.cell(9, c)
    if cell.fill.patternType == 'solid':
        gray_fills[c] = copy(cell.fill)

prev_code = None
is_gray = True  # 첫 단지 만나면 False(흰) 되도록
applied = 0
for r in range(5, 5 + len(rows)):
    code = ws.cell(r, 4).value
    if code is None: continue
    if code != prev_code:
        is_gray = not is_gray
        prev_code = code
    if is_gray:
        for col, fill in gray_fills.items():
            ws.cell(r, col).fill = copy(fill)
        applied += 1
print(f'회색 적용 행: {applied}')

wb.save(DST)
print(f'저장: {DST}')
