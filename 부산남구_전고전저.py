"""
시세시트_링크추가.xlsx '부산 남구' 탭에 전고/전저 채우기.
- 전고: 2021.01.01 ~ 2022.12.31 매매 최고가 → U열
- 전저: 2022.01.01 ~ 2023.12.31 매매 최저가 → V열
매칭키: (동, 건축년도, 반올림 전용면적) + 단지명 정규화 유사도
"""
import sys, io, os, re, csv, datetime
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from collections import defaultdict, Counter

BASE = r'C:\Users\한나\OneDrive\강성업무용\바탕 화면\문서\CLAUDE'
FILE = os.path.join(BASE, '시세시트_링크추가.xlsx')
DATA_CSV = os.path.join(BASE, '부산남구_시세', '매매_원본.csv')

HIGH_FROM = datetime.date(2021, 1, 1); HIGH_TO = datetime.date(2022, 12, 31)
LOW_FROM  = datetime.date(2022, 1, 1); LOW_TO  = datetime.date(2023, 12, 31)


def norm_name(s):
    if not s: return ''
    s = str(s).lower(); s = re.sub(r'\s+', '', s); s = re.sub(r'[()\[\]{}·.\-_,/&+]', '', s)
    return s


def norm_dong(s):
    if not s: return ''
    s = str(s).strip()
    m = re.search(r'^(.+?[읍면동])(\s|$)', s)
    return m.group(1) if m else s


def parse_amt(s):
    if s is None: return None
    s = str(s).replace(',', '').strip()
    if not s: return None
    try: return int(s)
    except:
        try: return int(float(s))
        except: return None


trades = []
with open(DATA_CSV, 'r', encoding='utf-8-sig') as f:
    for r in csv.DictReader(f):
        try:
            y = int(r['dealYear']); m = int(r['dealMonth']); d = int(r['dealDay'])
            dt = datetime.date(y, m, d)
            area = float(r['excluUseAr'])
            by = int(r['buildYear']) if r['buildYear'] else None
        except: continue
        amt = parse_amt(r.get('dealAmount'))
        if amt is None or by is None: continue
        trades.append({
            'date': dt, 'dong': norm_dong(r.get('umdNm')),
            'apt': (r.get('aptNm') or '').strip(),
            'apt_norm': norm_name(r.get('aptNm')),
            'build_year': by, 'area_round': round(area), 'amount': amt,
        })
print(f'매매 {len(trades)}건')

idx = defaultdict(list)
for it in trades:
    idx[(it['dong'], it['build_year'], it['area_round'])].append(it)


def best_match(dong, byear, area_round, name, tol=1):
    cand = []
    for da in range(-tol, tol+1):
        cand.extend(idx.get((dong, byear, area_round + da), []))
    if not cand: return [], 'none'
    wn = norm_name(name)
    exact = [c for c in cand if c['apt_norm'] == wn]
    if exact: return exact, 'exact'
    sub = [c for c in cand if wn and (wn in c['apt_norm'] or c['apt_norm'] in wn)]
    if sub:
        top = Counter(c['apt'] for c in sub).most_common(1)[0][0]
        return [c for c in sub if c['apt'] == top], 'substring'
    top = Counter(c['apt'] for c in cand).most_common(1)[0][0]
    return [c for c in cand if c['apt'] == top], 'fallback'


wb = openpyxl.load_workbook(FILE)
ws = wb['부산 남구']

COL = {'dong':1, 'name':5, 'year':6, 'area':11, 'high':21, 'low':22}
high_n = low_n = proc = 0
none_rows = []
for r in range(5, ws.max_row + 1):
    dong = ws.cell(r, COL['dong']).value
    name = ws.cell(r, COL['name']).value
    y = ws.cell(r, COL['year']).value
    a = ws.cell(r, COL['area']).value
    if not (dong and name and y and a): continue
    try:
        byear = int(float(y))
        ar = round(float(a))
    except: continue
    proc += 1
    items, flag = best_match(norm_dong(dong), byear, ar, str(name))
    if flag == 'none':
        none_rows.append((r, dong, name, byear, a))
        continue
    high = max((i['amount'] for i in items if HIGH_FROM <= i['date'] <= HIGH_TO), default=None)
    low = min((i['amount'] for i in items if LOW_FROM <= i['date'] <= LOW_TO), default=None)
    if high is not None: ws.cell(r, COL['high'], value=high); high_n += 1
    if low is not None: ws.cell(r, COL['low'], value=low); low_n += 1

wb.save(FILE)
print(f'처리 행: {proc}')
print(f'전고 채움: {high_n}')
print(f'전저 채움: {low_n}')
print(f'매칭 실패: {len(none_rows)}')
for r, d, n, y, a in none_rows[:10]:
    print(f'  행{r} {d} {n} {y} 전용{a}')
