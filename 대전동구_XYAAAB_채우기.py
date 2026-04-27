"""
대전동구 탭의 X, Y, AA, AB 채우기
- 월별 매매/전세 평균 → 전세가율 계산
- 전세가율 80% 이상 연속 구간 중, 최저 매매가가 발생한 구간 선택
- X: 최저매매가 발생 월 "YYYY.MM"
- Y: 해당 구간 길이 (Y.MM 형식, 11개월=0.11, 12개월=1.0, 15개월=1.03)
- AA: 최저매매가 (만원)
- AB: 그 월의 전세가 (만원)
"""
import sys, io, os, re, csv, datetime
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from collections import defaultdict, Counter
from statistics import mean

BASE = r'C:\Users\한나\OneDrive\강성업무용\바탕 화면\문서\CLAUDE'
FILE = os.path.join(BASE, '시세시트_링크추가.xlsx')
DATA = os.path.join(BASE, '대전동구_시세')

RATIO_THRESHOLD = 0.80


def norm_name(s):
    if not s: return ''
    return re.sub(r'[()\[\]{}·.\-_,/&+]', '', re.sub(r'\s+', '', str(s).lower()))

def norm_dong(s):
    if not s: return ''
    s = str(s).strip()
    m = re.search(r'^(.+?[읍면동])(\s|$)', s)
    return m.group(1) if m else s

def parse_amt(s):
    if s is None: return None
    s = str(s).replace(',', '').strip()
    if not s: return None
    try: return int(s)
    except:
        try: return int(float(s))
        except: return None


def load_csv(path):
    with open(path, 'r', encoding='utf-8-sig') as f:
        return list(csv.DictReader(f))


def prepare(rows, amt_key, is_rent=False):
    out = []
    for r in rows:
        try:
            y = int(r['dealYear']); m = int(r['dealMonth'])
            area = float(r['excluUseAr'])
            by = int(r['buildYear']) if r['buildYear'] else None
        except: continue
        amt = parse_amt(r.get(amt_key))
        if amt is None or by is None: continue
        # 전세는 신규만 (갱신 제외)
        if is_rent and (r.get('contractType') or '').strip() == '갱신':
            continue
        # 전세는 보증금만 (월세 큰 건 제외하려면 여기서 필터 가능하지만 일단 포함)
        out.append({
            'ym': (y, m), 'dong': norm_dong(r.get('umdNm')),
            'apt': (r.get('aptNm') or '').strip(),
            'apt_norm': norm_name(r.get('aptNm')),
            'build_year': by, 'area_round': round(area), 'amount': amt,
        })
    return out


print('CSV 로드...')
trades = prepare(load_csv(os.path.join(DATA, '매매_원본.csv')), 'dealAmount')
rents  = prepare(load_csv(os.path.join(DATA, '전세_원본.csv')), 'deposit', is_rent=True)
print(f'  매매 {len(trades)}건 / 전세 {len(rents)}건')


def build_idx(items):
    idx = defaultdict(list)
    for it in items:
        idx[(it['dong'], it['build_year'], it['area_round'])].append(it)
    return idx


trade_idx = build_idx(trades)
rent_idx = build_idx(rents)


def best_match(idx, dong, byear, area_round, name, tol=1):
    cand = []
    for da in range(-tol, tol+1):
        cand.extend(idx.get((dong, byear, area_round + da), []))
    if not cand: return []
    wn = norm_name(name)
    exact = [c for c in cand if c['apt_norm'] == wn]
    if exact: return exact
    sub = [c for c in cand if wn and (wn in c['apt_norm'] or c['apt_norm'] in wn)]
    if sub:
        top = Counter(c['apt'] for c in sub).most_common(1)[0][0]
        return [c for c in sub if c['apt'] == top]
    top = Counter(c['apt'] for c in cand).most_common(1)[0][0]
    return [c for c in cand if c['apt'] == top]


def monthly_items(items):
    """{(y,m): [amounts]}"""
    bucket = defaultdict(list)
    for it in items:
        bucket[it['ym']].append(it['amount'])
    return bucket


def ym_next(y, m):
    return (y, m+1) if m < 12 else (y+1, 1)


def ym_prev(y, m):
    return (y, m-1) if m > 1 else (y-1, 12)


def ym_shift(ym, delta):
    y, m = ym
    for _ in range(abs(delta)):
        if delta > 0: y, m = ym_next(y, m)
        else: y, m = ym_prev(y, m)
    return (y, m)


ROLL = 6  # 6개월 이동평균


def rolling_avg(bucket, months_range):
    """각 월에 대해 최근 ROLL개월(M-ROLL+1 ~ M)의 평균 반환"""
    out = {}
    for ym in months_range:
        vals = []
        for k in range(ROLL):
            vals.extend(bucket.get(ym_shift(ym, -k), []))
        if vals:
            out[ym] = mean(vals)
    return out


def find_best_period(trade_items, rent_items):
    """6개월 이동평균 전세가율 80%+ 연속 구간 중 min 매매가 run 반환"""
    t_bucket = monthly_items(trade_items)
    r_bucket = monthly_items(rent_items)
    all_yms = sorted(set(t_bucket.keys()) | set(r_bucket.keys()))
    if not all_yms: return None
    # 전체 월 범위 생성
    ymin, ymax = all_yms[0], all_yms[-1]
    months = []
    cur = ymin
    while cur <= ymax:
        months.append(cur)
        cur = ym_next(*cur)

    t_avg = rolling_avg(t_bucket, months)
    r_avg = rolling_avg(r_bucket, months)

    common = [ym for ym in months if ym in t_avg and ym in r_avg and t_avg[ym] > 0]
    if not common: return None
    ratios = {ym: r_avg[ym] / t_avg[ym] for ym in common}
    high = [ym for ym in sorted(ratios) if ratios[ym] >= RATIO_THRESHOLD]
    if not high: return None

    # 연속 그룹핑 (바로 이어지는 월만)
    runs = []
    cur_run = [high[0]]
    for ym in high[1:]:
        if ym_next(*cur_run[-1]) == ym:
            cur_run.append(ym)
        else:
            runs.append(cur_run); cur_run = [ym]
    runs.append(cur_run)

    best = None
    for run in runs:
        min_ym = min(run, key=lambda y: t_avg[y])
        min_t = t_avg[min_ym]
        if best is None or min_t < best[0]:
            best = (min_t, run, min_ym)

    min_t, run, min_ym = best
    rent_at_min = r_avg[min_ym]
    return (run[0], run[-1], min_ym, round(min_t), round(rent_at_min), len(run))


def months_to_display(n):
    """개월수 → Y.MM (예: 11→0.11, 12→1.0, 15→1.03, 24→2.0)"""
    years, extra = divmod(n, 12)
    # 0.11 또는 1.0 또는 1.03 같은 표기 → float
    return years + extra / 100.0


# 엑셀 처리: 값 읽기용(data_only)과 쓰기용 따로
wb_ro = openpyxl.load_workbook(FILE, data_only=True)
ws_ro = wb_ro['대전동구']
wb = openpyxl.load_workbook(FILE)
ws = wb['대전동구']

# 대전동구 컬럼: A=동, D=단지명, H=연식, L=전용면적, X=시기, Y=기간, AA=매매가, AB=전세가
COL = {'dong':1, 'name':4, 'year':8, 'area':12, 'X':24, 'Y':25, 'AA':27, 'AB':28}

proc = filled = 0
no_data = []

# 수식을 계산 값으로 돌려받기: data_only=True 시트에서 읽되,
# 수식 미계산(None)일 경우 raw 값 파싱
def read_val(ws_ro, ws_raw, r, c):
    v = ws_ro.cell(r, c).value
    if v is None:
        rv = ws_raw.cell(r, c).value
        return rv
    return v

# D열 수식 =D(이전행) 을 처리: 위로 올라가며 실제 단지명 찾기
def resolve_name(r):
    cur = r
    while cur >= 5:
        val = ws_ro.cell(cur, 4).value  # data_only 값
        if val is None:
            val = ws.cell(cur, 4).value
        if val is None: return None
        s = str(val)
        if s.startswith('='):
            # =D(n) 같은 참조면 그 행 가기
            m = re.match(r'=\s*D\s*(\d+)', s)
            if m:
                cur = int(m.group(1))
                continue
            return None
        return s
    return None

for r in range(5, ws.max_row + 1):
    dong = ws.cell(r, COL['dong']).value
    name = resolve_name(r)
    y_val = ws.cell(r, COL['year']).value
    a_val = ws.cell(r, COL['area']).value
    if not (dong and name and y_val and a_val): continue
    try:
        byear = int(float(y_val))
        ar = round(float(a_val))
    except: continue
    proc += 1

    t_items = best_match(trade_idx, norm_dong(dong), byear, ar, str(name))
    r_items = best_match(rent_idx, norm_dong(dong), byear, ar, str(name))
    if not t_items or not r_items:
        no_data.append((r, name, byear, a_val, '매매' if not t_items else '전세', '없음'))
        continue

    result = find_best_period(t_items, r_items)
    if result is None:
        no_data.append((r, name, byear, a_val, '전세가율 80%+ 기간', '없음'))
        continue
    start_ym, end_ym, min_ym, min_t, rent_at_min, length = result
    # X: "2022.03" 형식 - float (2022.03)
    ws.cell(r, COL['X'], value=min_ym[0] + min_ym[1]/100.0)
    ws.cell(r, COL['Y'], value=months_to_display(length))
    ws.cell(r, COL['AA'], value=min_t)
    ws.cell(r, COL['AB'], value=rent_at_min)
    filled += 1

wb.save(FILE)
print(f'처리 행: {proc}, 채움: {filled}, 데이터없음: {len(no_data)}')
for x in no_data[:15]:
    print(f'  행{x[0]} {x[1]} {x[2]} 전용{x[3]} — {x[4]} {x[5]}')
