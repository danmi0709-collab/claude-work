# -*- coding: utf-8 -*-
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from pathlib import Path
import openpyxl

path = Path(r'C:\Users\한나\OneDrive\월부\202605 실전 송파구\(송파구) 매임 체크리스트.xlsx')
wb = openpyxl.load_workbook(path)
print('시트:', wb.sheetnames)
ws = wb.active
print('크기:', ws.dimensions)
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True)):
    if any(v is not None for v in row):
        print(f"행{i+1}: {[v for v in row[:15]]}")
