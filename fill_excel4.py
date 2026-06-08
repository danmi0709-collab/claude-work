# -*- coding: utf-8 -*-
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from pathlib import Path
import openpyxl

path = Path(r'C:\Users\한나\OneDrive\월부\202605 실전 송파구\(송파구) 매임 체크리스트.xlsx')
wb = openpyxl.load_workbook(path)
ws = wb.active

# Q~S열 이미 입력. 이번에 T~W열 (나머지 4개)

properties = [
    # 상단 4번
    ('316동 1504호', 30.0,  '헬리오시티', '4베이판상형(F타입)', None,  None,       '4베이 판상형 F타입',           '공실'),
    # 하단 3개
    ('218동 3503호', 28.5,  '헬리오시티', '타워형(D타입)',       None,  '2027년5월', '원세 2억/400 27/5만기',        '세입자'),
    ('201동 703호',  29.5,  '헬리오시티', '타워형(D타입)',       None,  '9월만기',   '입주 9월만기',                  '공실'),
    ('515동 1503호', 31.0,  '헬리오시티', '타워형(D타입)',       11.3,  '2027년2월', '전세 11.3억 27/2만기',         '세입자'),
]

start_col = 20  # T열

for i, (dong_ho, price, apt, apt_type, jeonse, expire, memo, resident) in enumerate(properties):
    col = start_col + i
    col_letter = ws.cell(row=1, column=col).column_letter

    ws.cell(row=3,  column=col).value = apt
    ws.cell(row=5,  column=col).value = '헬리오시티부동산'
    ws.cell(row=11, column=col).value = dong_ho
    ws.cell(row=13, column=col).value = apt_type
    ws.cell(row=15, column=col).value = resident
    ws.cell(row=18, column=col).value = memo
    ws.cell(row=19, column=col).value = price
    if jeonse:
        ws.cell(row=21, column=col).value = jeonse
    if expire:
        ws.cell(row=17, column=col).value = expire

    print(f"  {col_letter}열 ({apt} {dong_ho}): 매매가 {price}억 입력 완료")

wb.save(path)
print("\n저장 완료!")
