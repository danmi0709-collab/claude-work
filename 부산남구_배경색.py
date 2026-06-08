"""
부산 남구 탭: 단지가 바뀔 때마다 회색/흰 토글.
첫 단지는 흰색(기본), 다음 단지는 회색, 이후 교대.
회색 패턴은 천안동남구 행9(휴먼시아) 기준 그대로 복제.
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from copy import copy

FILE = r'C:\Users\한나\OneDrive\강성업무용\바탕 화면\문서\CLAUDE\시세시트_링크추가.xlsx'

wb = openpyxl.load_workbook(FILE)

# 천안동남구 행9의 회색 패턴 추출
ref_ws = wb['천안동남구']
gray_fills = {}
for c in range(1, 52):
    cell = ref_ws.cell(9, c)
    if cell.fill.patternType == 'solid':
        gray_fills[c] = copy(cell.fill)
print(f'참조 회색 열 수: {len(gray_fills)}')

ws = wb['부산 남구']
START = 5
END = ws.max_row

prev_code = None
is_gray = True  # 첫 단지 처음 만나면 not-True=False(흰색). 아래에서 진입 시 토글하도록 True로 시작.

applied = 0
for r in range(START, END + 1):
    code = ws.cell(r, 4).value
    if code is None:
        continue
    if code != prev_code:
        is_gray = not is_gray
        prev_code = code
    if is_gray:
        for col, fill in gray_fills.items():
            ws.cell(r, col).fill = copy(fill)
        applied += 1

wb.save(FILE)
print(f'회색 적용 행: {applied}')
