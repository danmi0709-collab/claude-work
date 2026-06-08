# -*- coding: utf-8 -*-
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from pathlib import Path
import openpyxl

path = Path(r'C:\Users\한나\OneDrive\월부\202605 실전 송파구\(송파구) 매임 체크리스트.xlsx')
wb = openpyxl.load_workbook(path)
ws = wb.active

# 이전: C~P열 (물건1~14) 입력 완료
# 이번: Q(물건15)~S(물건17) - 헬리오시티 동그라미 3개

# (동호수, 매매가(억), 아파트명, 타입, 전세가(억), 만기일, 특이사항, 거주자)
properties = [
    ('404동 1002호', 31.0,  '헬리오시티', '타워형(D타입)', None,  None,       '타워형 D타입',                '공실'),
    ('402동 2204호', 30.0,  '헬리오시티', '4베이판상형(F타입)', 12.5, '2027년8월', '4베이 판상형 F타입 전세12.5억', '세입자'),
    ('219동 2602호', 30.0,  '헬리오시티', '판상형(A타입)',  None,  None,       '판상형 A타입',                '공실'),
]

start_col = 17  # Q열 = 물건15

for i, (dong_ho, price, apt, apt_type, jeonse, expire, memo, resident) in enumerate(properties):
    col = start_col + i
    col_letter = ws.cell(row=1, column=col).column_letter

    ws.cell(row=3,  column=col).value = apt
    ws.cell(row=5,  column=col).value = '헬리오시티부동산'
    ws.cell(row=11, column=col).value = dong_ho
    ws.cell(row=13, column=col).value = apt_type
    ws.cell(row=15, column=col).value = resident
    ws.cell(row=18, column=col).value = memo
    ws.cell(row=19, column=col).value = price
    if jeonse:
        ws.cell(row=21, column=col).value = jeonse
    if expire:
        ws.cell(row=17, column=col).value = expire

    print(f"  {col_letter}열 ({apt} {dong_ho}): 매매가 {price}억 입력 완료")

wb.save(path)
print("\n저장 완료!")
