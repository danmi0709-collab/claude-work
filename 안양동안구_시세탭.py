"""안양동안구 탭 생성 + 전고전저 채우기
사용:
  1단계(탭생성): python 안양동안구_시세탭.py tab
  2단계(데이터수집): python 안양동안구_시세탭.py fetch
  3단계(전고전저): python 안양동안구_시세탭.py high
  전체: python 안양동안구_시세탭.py all
"""
import sys, io, os, re, csv, json, time, datetime, urllib.parse, urllib.request
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, PatternFill
from collections import defaultdict, Counter

BASE = r'C:\Users\한나\OneDrive\강성업무용\바탕 화면\문서\CLAUDE'
FILE = os.path.join(BASE, '시세시트_링크추가.xlsx')
SRC  = os.path.join(BASE, '안양시 동안구.xlsx')
DATA = os.path.join(BASE, '안양동안구_시세')
CSV  = os.path.join(DATA, '매매_원본.csv')
CFG  = os.path.join(DATA, 'config.json')

URL_API = 'https://apis.data.go.kr/1613000/RTMSDataSvcAptTradeDev/getRTMSDataSvcAptTradeDev'
URL_NAVER = 'https://new.land.naver.com/complexes/{}?ms=2AfAQG,3zprE6,17&a=APT:ABYG:JGC&b=A1:B1&e=RETAIL&h=34&i=132&l=198&ad=true'
PYONG = 3.3058
GRAY  = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
WHITE = PatternFill(fill_type=None)

HIGH_FROM = datetime.date(2021, 1, 1);  HIGH_TO = datetime.date(2022, 12, 31)
LOW_FROM  = datetime.date(2022, 1, 1);  LOW_TO  = datetime.date(2023, 12, 31)

mode = sys.argv[1] if len(sys.argv) > 1 else 'all'

def extract_num(s):
    m = re.search(r'(\d+)', str(s)) if s else None
    return int(m.group(1)) if m else None

def norm_name(s):
    return re.sub(r'[()\[\]{}·.\-_,/&+]', '', re.sub(r'\s+', '', str(s).lower())) if s else ''

def norm_dong(s):
    if not s: return ''
    m = re.search(r'^(.+?[읍면동])(\s|$)', str(s).strip())
    return m.group(1) if m else str(s).strip()

def parse_amt(s):
    if s is None: return None
    s = str(s).replace(',', '').strip()
    try: return int(s)
    except:
        try: return int(float(s))
        except: return None

# ── 1. 안양시 동안구.xlsx 읽기 ──
print('=== 안양시 동안구.xlsx 로드 ===')
src_wb = openpyxl.load_workbook(SRC, data_only=True)
src_ws = src_wb.worksheets[0]
rows = []
for r in range(2, src_ws.max_row + 1):
    code = src_ws.cell(r, 5).value
    if code is None: continue
    try: code = int(code)
    except: continue
    rows.append({
        'dong':     src_ws.cell(r, 3).value,
        'code':     code,
        'name':     src_ws.cell(r, 6).value,
        'approved': src_ws.cell(r, 7).value,
        'hh_tot':   src_ws.cell(r, 8).value,
        'area_excl':src_ws.cell(r, 9).value,
        'struct':   src_ws.cell(r, 10).value,
        'rooms':    src_ws.cell(r, 11).value,
        'baths':    src_ws.cell(r, 12).value,
        'type_nm':  src_ws.cell(r, 13).value,
        'hh_type':  src_ws.cell(r, 14).value,
    })
print(f'  {len(rows)}행 로드')

# ── 2. 데이터 수집 (fetch / all) ──
if mode in ('fetch', 'all'):
    print('\n=== 매매 데이터 수집 ===')
    with open(CFG, encoding='utf-8') as f:
        cfg = json.load(f)
    if '여기에' in cfg.get('service_key', ''):
        print('⚠ config.json에 service_key를 먼저 입력하세요!')
        if mode == 'fetch': sys.exit(1)
    else:
        def fetch_all(ymd):
            out = []; page = 1
            while True:
                p = {'serviceKey': cfg['service_key'], 'LAWD_CD': cfg['lawd_cd'],
                     'DEAL_YMD': ymd, 'numOfRows': '1000', 'pageNo': str(page), '_type': 'json'}
                u = URL_API + '?' + urllib.parse.urlencode(p, safe=':/')
                with urllib.request.urlopen(u, timeout=60) as r:
                    data = json.loads(r.read().decode('utf-8'))
                body = data.get('response', {}).get('body', {})
                items = body.get('items', {})
                if not items: break
                lst = items.get('item', [])
                if isinstance(lst, dict): lst = [lst]
                out.extend(lst)
                if page * 1000 >= int(body.get('totalCount', 0)): break
                page += 1; time.sleep(0.3)
            return out

        def months_between(s, e):
            sy, sm = int(s[:4]), int(s[4:]); ey, em = int(e[:4]), int(e[4:])
            out = []; y, m = sy, sm
            while (y, m) <= (ey, em):
                out.append(f'{y:04d}{m:02d}'); m += 1
                if m > 12: m = 1; y += 1
            return out

        TF = ['dealYear','dealMonth','dealDay','umdNm','aptNm','buildYear',
              'excluUseAr','dealAmount','floor','jibun','aptSeq']
        ym_list = months_between('201801', '202604')
        all_items = []
        for i, ym in enumerate(ym_list, 1):
            try:
                it = fetch_all(ym)
                print(f'  [매매] {ym}: {len(it)}건 ({i}/{len(ym_list)})')
                all_items.extend(it); time.sleep(0.3)
            except Exception as e:
                print(f'  [매매] {ym}: 에러 {e}')
        with open(CSV, 'w', encoding='utf-8-sig', newline='') as f:
            w = csv.DictWriter(f, fieldnames=TF, extrasaction='ignore')
            w.writeheader()
            for it in all_items: w.writerow(it)
        print(f'  저장: {CSV} ({len(all_items)}건)')

# ── 3. 탭 생성 (tab / all) ──
if mode in ('tab', 'all'):
    print('\n=== 안양동안구 탭 생성 ===')
    wb = openpyxl.load_workbook(FILE)
    if '안양동안구' in wb.sheetnames:
        del wb['안양동안구']
    ws = wb.copy_worksheet(wb['창원성산구'])
    ws.title = '안양동안구'

    # 기존 데이터 초기화
    for r in range(5, ws.max_row + 1):
        for c in range(1, 42):
            ws.cell(r, c).value = None
            ws.cell(r, c).hyperlink = None
            ws.cell(r, c).fill = WHITE

    # 데이터 입력
    for i, d in enumerate(rows):
        r = 5 + i
        ws.cell(r, 1, value=d['dong'])
        ws.cell(r, 4, value=d['code'])
        nc = ws.cell(r, 5, value=d['name'])
        nc.hyperlink = URL_NAVER.format(d['code'])
        nc.font = Font(name=nc.font.name, size=nc.font.size, bold=nc.font.bold,
                       color='0563C1', underline='single')
        ws.cell(r, 6, value=d['approved'])
        ws.cell(r, 7, value=d['hh_tot'])
        ws.cell(r, 8, value=f'=2026-INT(F{r})+1')
        supply = extract_num(d['type_nm'])
        if supply:
            ws.cell(r, 10, value=supply)
            ws.cell(r, 12, value=round(supply / PYONG))
        try: ws.cell(r, 11, value=int(float(d['area_excl'])))
        except: ws.cell(r, 11, value=d['area_excl'])
        ws.cell(r, 13, value=d['struct'])
        ws.cell(r, 14, value=d['rooms'])
        ws.cell(r, 15, value=d['baths'])
        ws.cell(r, 16, value=d['hh_type'])

    # 배경색 토글 (단지별 회색/흰색 교차)
    prev_key = None; use_gray = False
    for i, d in enumerate(rows):
        r = 5 + i
        key = (d['dong'], d['name'], d['approved'])
        if key != prev_key:
            use_gray = not use_gray; prev_key = key
        if use_gray:
            for c in range(1, 38): ws.cell(r, c).fill = GRAY

    wb.save(FILE)
    print(f'  {len(rows)}행 입력 + 배경색 완료')

# ── 4. 전고전저 (high / all) ──
if mode in ('high', 'all'):
    print('\n=== 전고전저 채우기 ===')
    if not os.path.exists(CSV):
        print(f'  ⚠ CSV 없음: {CSV}')
        print('  먼저 fetch 모드로 데이터 수집 필요')
    else:
        trades = []
        with open(CSV, 'r', encoding='utf-8-sig') as f:
            for row in csv.DictReader(f):
                try:
                    y=int(row['dealYear']); m=int(row['dealMonth']); d=int(row['dealDay'])
                    dt=datetime.date(y,m,d); area=float(row['excluUseAr'])
                    by=int(row['buildYear']) if row['buildYear'] else None
                except: continue
                amt = parse_amt(row.get('dealAmount'))
                if amt is None or by is None: continue
                trades.append({'date': dt, 'dong': norm_dong(row.get('umdNm')),
                    'apt': (row.get('aptNm') or '').strip(),
                    'apt_norm': norm_name(row.get('aptNm')),
                    'build_year': by, 'area_round': round(area), 'amount': amt})
        print(f'  매매 {len(trades)}건')

        idx = defaultdict(list)
        for it in trades: idx[(it['dong'], it['build_year'], it['area_round'])].append(it)

        def best_match(dong, byear, ar, name, tol=1):
            cand = []
            for da in range(-tol, tol+1): cand.extend(idx.get((dong, byear, ar+da), []))
            if not cand: return []
            wn = norm_name(name)
            exact = [c for c in cand if c['apt_norm'] == wn]
            if exact: return exact
            sub = [c for c in cand if wn and (wn in c['apt_norm'] or c['apt_norm'] in wn)]
            if sub:
                top = Counter(c['apt'] for c in sub).most_common(1)[0][0]
                return [c for c in sub if c['apt'] == top]
            top = Counter(c['apt'] for c in cand).most_common(1)[0][0]
            return [c for c in cand if c['apt'] == top]

        wb = openpyxl.load_workbook(FILE)
        ws = wb['안양동안구']
        high_n = low_n = proc = 0
        for r in range(5, ws.max_row + 1):
            dong  = ws.cell(r, 1).value;  name = ws.cell(r, 5).value
            y_v   = ws.cell(r, 6).value;  a_v  = ws.cell(r, 11).value
            if not (dong and name and y_v and a_v): continue
            try: byear = int(float(str(y_v)[:4])); ar = round(float(a_v))
            except: continue
            proc += 1
            items = best_match(norm_dong(str(dong)), byear, ar, str(name))
            if not items: continue
            high = max((i['amount'] for i in items if HIGH_FROM <= i['date'] <= HIGH_TO), default=None)
            low  = min((i['amount'] for i in items if LOW_FROM  <= i['date'] <= LOW_TO),  default=None)
            if high is not None: ws.cell(r, 21, value=round(high/10000, 2)); high_n += 1
            if low  is not None: ws.cell(r, 22, value=round(low /10000, 2)); low_n  += 1

        wb.save(FILE)
        print(f'  처리 {proc}행 | 전고 {high_n}행 | 전저 {low_n}행')

print('\n완료!')
