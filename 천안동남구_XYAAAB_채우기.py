"""
천안동남구 탭의 AA/AB/AD/AE 채우기 (대전동구와 같은 방식)
- 월별 매매/전세 6개월 이동평균 → 전세가율
- 80%+ 연속 구간 중 최저매매가 발생한 구간 선택
- AA(시기): '2020.03' 형식 / AB(기간): 0.11=11달, 1.0=1년
- AD(매매가): 최저매매 / AE(전세가): 그 월의 전세
"""
import sys, io, os, re, csv, datetime
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from collections import defaultdict, Counter
from statistics import mean

BASE = r'C:\Users\한나\OneDrive\강성업무용\바탕 화면\문서\CLAUDE'
FILE = os.path.join(BASE, '시세시트_링크추가.xlsx')
DATA = os.path.join(BASE, '천안동남구_시세')

RATIO_THRESHOLD = 0.80
ROLL = 6


def norm_name(s):
    if not s: return ''
    return re.sub(r'[()\[\]{}·.\-_,/&+]', '', re.sub(r'\s+', '', str(s).lower()))

def norm_dong(s):
    if not s: return ''
    s = str(s).strip()
    m = re.search(r'^(.+?[읍면동])(\s|$)', s)
    return m.group(1) if m else s

def parse_amt(s):
    if s is None: return None
    s = str(s).replace(',', '').strip()
    if not s: return None
    try: return int(s)
    except:
        try: return int(float(s))
        except: return None


def load_csv(p):
    with open(p, 'r', encoding='utf-8-sig') as f:
        return list(csv.DictReader(f))


def prepare(rows, amt_key, is_rent=False):
    out = []
    for r in rows:
        try:
            y = int(r['dealYear']); m = int(r['dealMonth'])
            area = float(r['excluUseAr'])
            by = int(r['buildYear']) if r['buildYear'] else None
        except: continue
        amt = parse_amt(r.get(amt_key))
        if amt is None or by is None: continue
        if is_rent and (r.get('contractType') or '').strip() == '갱신': continue
        out.append({
            'ym': (y, m), 'dong': norm_dong(r.get('umdNm')),
            'apt': (r.get('aptNm') or '').strip(),
            'apt_norm': norm_name(r.get('aptNm')),
            'build_year': by, 'area_round': round(area), 'amount': amt,
        })
    return out


print('CSV 로드...')
trades = prepare(load_csv(os.path.join(DATA, '매매_원본.csv')), 'dealAmount')
rents  = prepare(load_csv(os.path.join(DATA, '전세_원본.csv')), 'deposit', is_rent=True)
print(f'  매매 {len(trades)}건 / 전세 {len(rents)}건')


def build_idx(items):
    idx = defaultdict(list)
    for it in items:
        idx[(it['dong'], it['build_year'], it['area_round'])].append(it)
    return idx


trade_idx = build_idx(trades)
rent_idx = build_idx(rents)


def best_match(idx, dong, byear, area_round, name, tol=1):
    cand = []
    for da in range(-tol, tol+1):
        cand.extend(idx.get((dong, byear, area_round + da), []))
    if not cand: return []
    wn = norm_name(name)
    exact = [c for c in cand if c['apt_norm'] == wn]
    if exact: return exact
    sub = [c for c in cand if wn and (wn in c['apt_norm'] or c['apt_norm'] in wn)]
    if sub:
        top = Counter(c['apt'] for c in sub).most_common(1)[0][0]
        return [c for c in sub if c['apt'] == top]
    top = Counter(c['apt'] for c in cand).most_common(1)[0][0]
    return [c for c in cand if c['apt'] == top]


def monthly_items(items):
    bucket = defaultdict(list)
    for it in items:
        bucket[it['ym']].append(it['amount'])
    return bucket


def ym_next(y, m): return (y, m+1) if m < 12 else (y+1, 1)
def ym_prev(y, m): return (y, m-1) if m > 1 else (y-1, 12)
def ym_shift(ym, delta):
    y, m = ym
    for _ in range(abs(delta)):
        if delta > 0: y, m = ym_next(y, m)
        else: y, m = ym_prev(y, m)
    return (y, m)


def rolling_avg(bucket, months):
    out = {}
    for ym in months:
        vals = []
        for k in range(ROLL):
            vals.extend(bucket.get(ym_shift(ym, -k), []))
        if vals: out[ym] = mean(vals)
    return out


def find_best_period(t_items, r_items):
    tb = monthly_items(t_items); rb = monthly_items(r_items)
    allyms = sorted(set(tb) | set(rb))
    if not allyms: return None
    months = []
    cur = allyms[0]
    while cur <= allyms[-1]:
        months.append(cur); cur = ym_next(*cur)
    t_avg = rolling_avg(tb, months)
    r_avg = rolling_avg(rb, months)
    common = [ym for ym in months if ym in t_avg and ym in r_avg and t_avg[ym] > 0]
    if not common: return None
    ratios = {ym: r_avg[ym] / t_avg[ym] for ym in common}
    high = [ym for ym in sorted(ratios) if ratios[ym] >= RATIO_THRESHOLD]
    if not high: return None
    runs = []; cur_run = [high[0]]
    for ym in high[1:]:
        if ym_next(*cur_run[-1]) == ym: cur_run.append(ym)
        else: runs.append(cur_run); cur_run = [ym]
    runs.append(cur_run)
    best = None
    for run in runs:
        min_ym = min(run, key=lambda y: t_avg[y])
        min_t = t_avg[min_ym]
        if best is None or min_t < best[0]: best = (min_t, run, min_ym)
    min_t, run, min_ym = best
    return (run[0], run[-1], min_ym, round(min_t), round(r_avg[min_ym]), len(run))


def months_to_display(n):
    years, extra = divmod(n, 12)
    return years + extra / 100.0


# ==== 엑셀 처리 (천안동남구) ====
wb_ro = openpyxl.load_workbook(FILE, data_only=True)
ws_ro = wb_ro['천안동남구']
wb = openpyxl.load_workbook(FILE)
ws = wb['천안동남구']

# 천안동남구: A(동)=1, E(단지명)=5, F(사용승인일)=6, K(전용면적)=11
# 출력: AA(시기)=27, AB(기간)=28, AD(매매가)=30, AE(전세가)=31
COL = {'dong':1, 'name':5, 'year':6, 'area':11,
       'time':27, 'period':28, 'trade':30, 'rent':31}


def resolve_name(r):
    cur = r
    while cur >= 5:
        val = ws_ro.cell(cur, COL['name']).value
        if val is None: val = ws.cell(cur, COL['name']).value
        if val is None: return None
        s = str(val)
        if s.startswith('='):
            m = re.match(r'=\s*[A-Z]+\s*(\d+)', s)
            if m: cur = int(m.group(1)); continue
            return None
        return s
    return None


proc = filled = 0
no_data = []
for r in range(5, ws.max_row + 1):
    dong = ws.cell(r, COL['dong']).value
    name = resolve_name(r)
    y_val = ws.cell(r, COL['year']).value
    a_val = ws.cell(r, COL['area']).value
    if not (dong and name and y_val and a_val): continue
    try:
        byear = int(float(y_val)); ar = round(float(a_val))
    except: continue
    proc += 1
    t = best_match(trade_idx, norm_dong(dong), byear, ar, str(name))
    rn = best_match(rent_idx, norm_dong(dong), byear, ar, str(name))
    if not t or not rn:
        no_data.append((r, name, byear, a_val, '매매' if not t else '전세', '없음')); continue
    res = find_best_period(t, rn)
    if res is None:
        no_data.append((r, name, byear, a_val, '전세가율 80%+', '없음')); continue
    _, _, min_ym, min_t, rent_at, length = res
    ws.cell(r, COL['time'], value=min_ym[0] + min_ym[1]/100.0)
    ws.cell(r, COL['period'], value=months_to_display(length))
    ws.cell(r, COL['trade'], value=min_t)
    ws.cell(r, COL['rent'], value=rent_at)
    filled += 1

wb.save(FILE)
print(f'처리 행: {proc}, 채움: {filled}, 데이터없음: {len(no_data)}')
