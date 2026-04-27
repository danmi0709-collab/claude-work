"""대전동구 미매칭 단지 수동 매핑으로 하이퍼링크 추가"""
import sys, io, os, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from openpyxl.styles import Font

BASE = r'C:\Users\한나\OneDrive\강성업무용\바탕 화면\문서\CLAUDE'
FILE = os.path.join(BASE, '시세시트_링크추가.xlsx')
REF = os.path.join(BASE, '동구 (1).xlsx')

URL = 'https://new.land.naver.com/complexes/{}?ms=2AfAQG,3zprE6,17&a=APT:ABYG:JGC&b=A1:B1&e=RETAIL&h=34&i=132&l=198&ad=true'

# 시트 단지명 → 동구(1) 단지명 수동 매핑
MANUAL = {
    '삼정그린코아포레1': '삼정그린코아포레스트1단지',
    '삼정그린코아포레2': '삼정그린코아포레스트2단지',
    '힐스테가양더와이즈(주복)': '힐스테이트가양더와이즈(주상복합)',
}

def norm(s):
    return re.sub(r'[()\[\]{}·.\-_,/&+\s]', '', str(s).lower()) if s else ''

# 참조 단지명→단지코드
wb_ref = openpyxl.load_workbook(REF, data_only=True)
ws_ref = wb_ref.worksheets[0]
name_to_code = {}
for r in range(2, ws_ref.max_row+1):
    code = ws_ref.cell(r, 5).value
    name = ws_ref.cell(r, 6).value
    if code and name:
        try: name_to_code[norm(name)] = int(code)
        except: pass

# 수동 매핑을 코드로 변환
manual_code = {}
for sheet_name, ref_name in MANUAL.items():
    c = name_to_code.get(norm(ref_name))
    if c:
        manual_code[norm(sheet_name)] = c
        print(f'  {sheet_name} → {ref_name} (코드: {c})')

# 시세시트 적용
wb = openpyxl.load_workbook(FILE)
ws = wb['대전동구']

added = 0
for r in range(5, ws.max_row+1):
    cell = ws.cell(r, 4)
    if cell.hyperlink: continue  # 이미 링크 있음
    name = cell.value
    if not name: continue
    code = manual_code.get(norm(name))
    if not code: continue
    cell.hyperlink = URL.format(code)
    ef = cell.font
    cell.font = Font(name=ef.name, size=ef.size, bold=ef.bold,
                     color='0563C1', underline='single')
    added += 1

wb.save(FILE)
print(f'\n추가 링크: {added}개')
