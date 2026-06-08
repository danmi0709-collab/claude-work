"""
송파구 탭을 시세시트_링크추가.xlsx에 추가
- 데이터 소스: 송파구/송파구 루시퍼홍.xlsx
- 창원성산구 탭과 동일한 형식
- 단지명 하이퍼링크 + 단지 토글 배경색 포함
"""
import sys, io, os, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, PatternFill
from copy import copy

BASE = r'C:\Users\한나\OneDrive\강성업무용\바탕 화면\문서\CLAUDE'
FILE = os.path.join(BASE, '시세시트_링크추가.xlsx')
SRC  = os.path.join(BASE, '송파구', '송파구 루시퍼홍.xlsx')

URL = 'https://new.land.naver.com/complexes/{}?ms=2AfAQG,3zprE6,17&a=APT:ABYG:JGC&b=A1:B1&e=RETAIL&h=34&i=132&l=198&ad=true'
PYONG = 3.3058

GRAY  = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
WHITE = PatternFill(fill_type=None)


def extract_pyeong_num(s):
    """'80C' '81A, 81B' → 80 처럼 첫 숫자 추출"""
    if s is None: return None
    m = re.search(r'(\d+)', str(s))
    return int(m.group(1)) if m else None


# ── 1. 원본 데이터 읽기 ──
print('루시퍼홍 파일 로드...')
src_wb = openpyxl.load_workbook(SRC, data_only=True)
src_ws = src_wb.worksheets[0]

rows = []
for r in range(2, src_ws.max_row + 1):
    code = src_ws.cell(r, 5).value
    if code is None: continue
    try: code = int(code)
    except: continue
    rows.append({
        'dong':      src_ws.cell(r, 3).value,   # 읍면동
        'code':      code,                        # 단지코드
        'name':      src_ws.cell(r, 6).value,    # 단지명
        'approved':  src_ws.cell(r, 7).value,    # 사용승인일
        'households':src_ws.cell(r, 8).value,    # 총세대수
        'area_excl': src_ws.cell(r, 9).value,    # 전용면적 (K)
        'struct':    src_ws.cell(r, 10).value,   # 구조 (M)
        'rooms':     src_ws.cell(r, 11).value,   # 방수 (N)
        'baths':     src_ws.cell(r, 12).value,   # 욕실수 (O)
        'type_name': src_ws.cell(r, 13).value,   # 평형명 → J(공급면적) 추출
        'hh_type':   src_ws.cell(r, 14).value,   # 평형별 세대수 (P)
    })
print(f'  {len(rows)}개 행 로드')


# ── 2. 대상 파일 열고 송파구 탭 생성 ──
wb = openpyxl.load_workbook(FILE)

# 기존 탭 있으면 삭제
if '송파구' in wb.sheetnames:
    del wb['송파구']

# 창원성산구 탭에서 헤더 복사하여 새 탭 생성
tmpl = wb['창원성산구']
ws = wb.copy_worksheet(tmpl)
ws.title = '송파구'

# 5행부터 기존 데이터 지우기
for r in range(5, ws.max_row + 1):
    for c in range(1, 42):
        cell = ws.cell(r, c)
        cell.value = None
        cell.hyperlink = None
        cell.fill = WHITE

print('송파구 탭 생성 완료')


# ── 3. 데이터 입력 ──
for i, d in enumerate(rows):
    r = 5 + i

    ws.cell(r, 1, value=d['dong'])       # A 동
    ws.cell(r, 4, value=d['code'])       # D 단지코드

    # E 단지명 + 하이퍼링크
    c_name = ws.cell(r, 5, value=d['name'])
    c_name.hyperlink = URL.format(d['code'])
    base = c_name.font
    c_name.font = Font(name=base.name, size=base.size, bold=base.bold,
                       color='0563C1', underline='single')

    ws.cell(r, 6, value=d['approved'])                     # F 사용승인일
    ws.cell(r, 7, value=d['households'])                   # G 총세대수
    ws.cell(r, 8, value=f'=2026-INT(F{r})+1')             # H 나이 수식

    # J 공급면적 (평형명에서 추출)
    supply = extract_pyeong_num(d['type_name'])
    if supply:
        ws.cell(r, 10, value=supply)
        ws.cell(r, 12, value=round(supply / PYONG))         # L 평형

    # K 전용면적
    try: ws.cell(r, 11, value=int(float(d['area_excl'])))
    except: ws.cell(r, 11, value=d['area_excl'])

    ws.cell(r, 13, value=d['struct'])                      # M 구조
    ws.cell(r, 14, value=d['rooms'])                       # N 방수
    ws.cell(r, 15, value=d['baths'])                       # O 욕실수
    ws.cell(r, 16, value=d['hh_type'])                     # P 평형별 세대수

print(f'  {len(rows)}행 데이터 입력 완료')


# ── 4. 단지 토글 배경색 ──
prev_key = None
use_gray = False
for i, d in enumerate(rows):
    r = 5 + i
    key = (d['dong'], d['name'], d['approved'])
    if key != prev_key:
        use_gray = not use_gray
        prev_key = key
    if use_gray:
        for c in range(1, 38):
            ws.cell(r, c).fill = GRAY

print('  배경색 토글 적용 완료')

wb.save(FILE)
print(f'\n저장 완료: {FILE}')
print(f'총 {len(rows)}행 입력')
