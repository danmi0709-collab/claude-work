"""
시세시트 파일의 천안동남구 시트에서, D열 단지코드를 이용해
E열 단지명 셀에 네이버 부동산 단지 페이지 하이퍼링크를 추가.
"""
import sys, io, os, shutil
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from openpyxl.styles import Font

BASE_DIR = r'C:\Users\한나\OneDrive\강성업무용\바탕 화면\문서\CLAUDE'
# \r 포함된 원본 파일명 찾기
src = None
for f in os.listdir(BASE_DIR):
    if '뽀오' in f and f.endswith('.xlsx'):
        src = f
        break
src_path = os.path.join(BASE_DIR, src)
backup = os.path.join(BASE_DIR, '시세시트_원본백업.xlsx')
out = os.path.join(BASE_DIR, '시세시트_링크추가.xlsx')

# 백업
shutil.copy(src_path, backup)
print(f'백업: {backup}')

URL_TEMPLATE = 'https://new.land.naver.com/complexes/{}?ms=2AfAQG,3zprE6,17&a=APT:ABYG:JGC&b=A1:B1&e=RETAIL&h=34&i=132&l=198&ad=true'

wb = openpyxl.load_workbook(src_path)
# 어느 시트에 단지코드/단지명 구조가 있는지 전 시트 검사
SHEET_TARGETS = []
for name in wb.sheetnames:
    ws = wb[name]
    # D4='단지코드', E4='단지명' 있는지 확인
    d4 = ws.cell(4, 4).value
    e4 = ws.cell(4, 5).value
    if str(d4).strip() == '단지코드' and str(e4).strip() == '단지명':
        SHEET_TARGETS.append(name)
print('처리 시트:', SHEET_TARGETS)

total_added = 0
for sheet_name in SHEET_TARGETS:
    ws = wb[sheet_name]
    added = 0
    seen = set()  # 같은 단지코드 여러 행에도 모두 링크
    for row in range(5, ws.max_row + 1):
        code_cell = ws.cell(row, 4)  # D
        name_cell = ws.cell(row, 5)  # E
        code = code_cell.value
        if code is None or not name_cell.value:
            continue
        try:
            code_int = int(code)
        except Exception:
            continue
        url = URL_TEMPLATE.format(code_int)
        name_cell.hyperlink = url
        # 링크 스타일 (파란색 + 밑줄)
        existing_font = name_cell.font
        name_cell.font = Font(
            name=existing_font.name,
            size=existing_font.size,
            bold=existing_font.bold,
            color='0563C1',
            underline='single',
        )
        added += 1
    print(f'  [{sheet_name}] {added}개 링크 추가')
    total_added += added

wb.save(out)
print(f'\n저장: {out}')
print(f'총 {total_added}개 링크 추가')
