"""창원성산구 탭 생성 + 데이터 채우기 + 하이퍼링크 + 배경색 토글"""
import sys, io, os, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.cell import MergedCell
from copy import copy

BASE = r'C:\Users\한나\OneDrive\강성업무용\바탕 화면\문서\CLAUDE'
SRC = os.path.join(BASE, '창원시 성산구.xlsx')
DST = os.path.join(BASE, '시세시트_링크추가.xlsx')

URL = 'https://new.land.naver.com/complexes/{}?ms=2AfAQG,3zprE6,17&a=APT:ABYG:JGC&b=A1:B1&e=RETAIL&h=34&i=132&l=198&ad=true'
PYONG = 3.3058

GRAY = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

def get_supply_area(s):
    """'59A, 59B' → 59, '112' → 112"""
    if s is None: return None
    m = re.search(r'\d+', str(s))
    return int(m.group()) if m else None


# ── 1. 창원시 성산구.xlsx 읽기 ──
print('창원시 성산구.xlsx 읽는 중...')
wb_src = openpyxl.load_workbook(SRC, data_only=True)
ws_src = wb_src.worksheets[0]

rows = []
for r in range(2, ws_src.max_row + 1):
    code = ws_src.cell(r, 5).value
    name = ws_src.cell(r, 6).value
    if not (code and name): continue
    rows.append({
        '동':       ws_src.cell(r, 3).value,
        '단지코드':  code,
        '단지명':    name,
        '사용승인일': ws_src.cell(r, 7).value,
        '총세대수':  ws_src.cell(r, 8).value,
        '전용면적':  ws_src.cell(r, 9).value,
        '구조':     ws_src.cell(r, 10).value,
        '방수':     ws_src.cell(r, 11).value,
        '욕실수':   ws_src.cell(r, 12).value,
        '공급명':   ws_src.cell(r, 13).value,
        '평형별세대수': ws_src.cell(r, 14).value,
    })
print(f'  {len(rows)}행 로드')


# ── 2. 시세시트에 탭 생성 ──
print('\n시세시트에 창원성산구 탭 생성...')
wb = openpyxl.load_workbook(DST)

# 이미 있으면 삭제 후 재생성
if '창원성산구' in wb.sheetnames:
    del wb['창원성산구']

# 부산 남구 탭에서 헤더(1~4행) 복사
ws_ref = wb['부산 남구']
ws = wb.create_sheet('창원성산구')

for r in range(1, 5):
    for c in range(1, ws_ref.max_column + 1):
        src_cell = ws_ref.cell(r, c)
        if isinstance(src_cell, MergedCell): continue
        dst_cell = ws.cell(r, c)
        dst_cell.value = src_cell.value
        if src_cell.has_style:
            dst_cell.font = copy(src_cell.font)
            dst_cell.fill = copy(src_cell.fill)
            dst_cell.border = copy(src_cell.border)
            dst_cell.alignment = copy(src_cell.alignment)
            dst_cell.number_format = src_cell.number_format

# 병합 셀 (헤더 1~4행만)
for mr in ws_ref.merged_cells.ranges:
    if mr.max_row <= 4:
        try: ws.merge_cells(str(mr))
        except: pass

# 열 너비
from openpyxl.utils import get_column_letter
for c in range(1, ws_ref.max_column + 1):
    cl = get_column_letter(c)
    if ws_ref.column_dimensions[cl].width:
        ws.column_dimensions[cl].width = ws_ref.column_dimensions[cl].width


# ── 3. 데이터 채우기 + 하이퍼링크 ──
print('데이터 채우기...')
prev_key = None
use_gray = False

for i, d in enumerate(rows):
    r = 5 + i

    # 배경색 토글 (단지 변경 시)
    key = (d['동'], d['단지코드'])
    if key != prev_key:
        use_gray = not use_gray
        prev_key = key

    fill = GRAY if use_gray else PatternFill(fill_type=None)

    ws.cell(r, 1, value=d['동'])         # A 동
    ws.cell(r, 4, value=d['단지코드'])   # D 단지코드

    # E 단지명 + 하이퍼링크
    name_cell = ws.cell(r, 5, value=d['단지명'])
    try:
        name_cell.hyperlink = URL.format(int(d['단지코드']))
        name_cell.font = Font(color='0563C1', underline='single')
    except: pass

    ws.cell(r, 6, value=d['사용승인일'])  # F
    ws.cell(r, 7, value=d['총세대수'])   # G
    ws.cell(r, 8, value=f'=2026-F{r}+1') # H 나이

    # J 공급면적, L 평형
    sup = get_supply_area(d['공급명'])
    if sup:
        ws.cell(r, 10, value=sup)
        ws.cell(r, 12, value=round(sup / PYONG))

    # K 전용면적
    try: ws.cell(r, 11, value=int(float(d['전용면적'])))
    except: ws.cell(r, 11, value=d['전용면적'])

    ws.cell(r, 13, value=d['구조'])       # M
    ws.cell(r, 14, value=d['방수'])       # N
    ws.cell(r, 15, value=d['욕실수'])     # O
    ws.cell(r, 16, value=d['평형별세대수']) # P

    # 배경색 적용
    for c in range(1, 39):
        ws.cell(r, c).fill = fill

wb.save(DST)
print(f'완료: 창원성산구 탭 {len(rows)}행 입력')
