"""부산 수영구: 매매 실거래가 수집(2021-01~2023-12) 후 전고/전저 채우기"""
import sys, io, os, json, time, csv, re, datetime, urllib.parse, urllib.request
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from collections import defaultdict, Counter

BASE_SELF = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(BASE_SELF)
FILE = os.path.join(ROOT, '시세시트_링크추가.xlsx')

with open(os.path.join(BASE_SELF, 'config.json'), 'r', encoding='utf-8') as f:
    cfg = json.load(f)

URL_TRADE = 'https://apis.data.go.kr/1613000/RTMSDataSvcAptTradeDev/getRTMSDataSvcAptTradeDev'


def fetch_all_month(ymd):
    out = []; page = 1
    while True:
        params = {'serviceKey': cfg['service_key'], 'LAWD_CD': cfg['lawd_cd'], 'DEAL_YMD': ymd,
                  'numOfRows': '1000', 'pageNo': str(page), '_type': 'json'}
        url = URL_TRADE + '?' + urllib.parse.urlencode(params, safe=':/')
        with urllib.request.urlopen(url, timeout=60) as r:
            data = json.loads(r.read().decode('utf-8'))
        body = data.get('response', {}).get('body', {})
        items = body.get('items', {})
        if not items: break
        lst = items.get('item', [])
        if isinstance(lst, dict): lst = [lst]
        out.extend(lst)
        total = int(body.get('totalCount', 0))
        if page * 1000 >= total: break
        page += 1; time.sleep(0.3)
    return out


def months_between(s, e):
    sy, sm = int(s[:4]), int(s[4:]); ey, em = int(e[:4]), int(e[4:])
    out = []; y, m = sy, sm
    while (y, m) <= (ey, em):
        out.append(f'{y:04d}{m:02d}'); m += 1
        if m > 12: m = 1; y += 1
    return out


CSV_PATH = os.path.join(BASE_SELF, '매매_원본.csv')
if not os.path.exists(CSV_PATH):
    print('실거래가 수집 중...')
    all_items = []
    ym_list = months_between('202101', '202312')
    for i, ym in enumerate(ym_list, 1):
        try:
            it = fetch_all_month(ym)
            print(f'  {ym}: {len(it)}건 ({i}/{len(ym_list)})')
            all_items.extend(it); time.sleep(0.3)
        except Exception as e:
            print(f'  {ym}: 에러 {e}')
    FIELDS = ['dealYear','dealMonth','dealDay','umdNm','aptNm','buildYear',
              'excluUseAr','dealAmount','floor','jibun','aptSeq']
    with open(CSV_PATH, 'w', encoding='utf-8-sig', newline='') as f:
        w = csv.DictWriter(f, fieldnames=FIELDS, extrasaction='ignore')
        w.writeheader()
        for it in all_items: w.writerow(it)
    print(f'저장: {CSV_PATH} ({len(all_items)}건)')


# 전고/전저 채우기
def norm_name(s):
    if not s: return ''
    return re.sub(r'[()\[\]{}·.\-_,/&+]', '', re.sub(r'\s+', '', str(s).lower()))

def norm_dong(s):
    if not s: return ''
    s = str(s).strip()
    m = re.search(r'^(.+?[읍면동])(\s|$)', s)
    return m.group(1) if m else s

def parse_amt(s):
    if s is None: return None
    s = str(s).replace(',', '').strip()
    if not s: return None
    try: return int(s)
    except:
        try: return int(float(s))
        except: return None


HIGH_FROM = datetime.date(2021, 1, 1); HIGH_TO = datetime.date(2022, 12, 31)
LOW_FROM  = datetime.date(2022, 1, 1); LOW_TO  = datetime.date(2023, 12, 31)

trades = []
with open(CSV_PATH, 'r', encoding='utf-8-sig') as f:
    for r in csv.DictReader(f):
        try:
            dt = datetime.date(int(r['dealYear']), int(r['dealMonth']), int(r['dealDay']))
            area = float(r['excluUseAr'])
            by = int(r['buildYear']) if r['buildYear'] else None
        except: continue
        amt = parse_amt(r.get('dealAmount'))
        if amt is None or by is None: continue
        trades.append({'date': dt, 'dong': norm_dong(r.get('umdNm')),
                       'apt': (r.get('aptNm') or '').strip(),
                       'apt_norm': norm_name(r.get('aptNm')),
                       'build_year': by, 'area_round': round(area), 'amount': amt})

idx = defaultdict(list)
for it in trades:
    idx[(it['dong'], it['build_year'], it['area_round'])].append(it)
print(f'매매 {len(trades)}건 로드')


def best_match(dong, byear, area_round, name, tol=1):
    cand = []
    for da in range(-tol, tol+1):
        cand.extend(idx.get((dong, byear, area_round + da), []))
    if not cand: return [], 'none'
    wn = norm_name(name)
    exact = [c for c in cand if c['apt_norm'] == wn]
    if exact: return exact, 'exact'
    sub = [c for c in cand if wn and (wn in c['apt_norm'] or c['apt_norm'] in wn)]
    if sub:
        top = Counter(c['apt'] for c in sub).most_common(1)[0][0]
        return [c for c in sub if c['apt'] == top], 'substring'
    top = Counter(c['apt'] for c in cand).most_common(1)[0][0]
    return [c for c in cand if c['apt'] == top], 'fallback'


wb = openpyxl.load_workbook(FILE)
ws = wb['부산수영구']
proc = high_n = low_n = 0
fails = []
for r in range(5, ws.max_row + 1):
    dong = ws.cell(r, 1).value; name = ws.cell(r, 5).value
    y = ws.cell(r, 6).value; a = ws.cell(r, 11).value
    if not (dong and name and y and a): continue
    try:
        byear = int(float(y)); ar = round(float(a))
    except: continue
    proc += 1
    items, flag = best_match(norm_dong(dong), byear, ar, str(name))
    if flag == 'none':
        fails.append((r, dong, name, byear, a)); continue
    high = max((i['amount'] for i in items if HIGH_FROM <= i['date'] <= HIGH_TO), default=None)
    low = min((i['amount'] for i in items if LOW_FROM <= i['date'] <= LOW_TO), default=None)
    if high is not None: ws.cell(r, 21, value=high); high_n += 1
    if low is not None: ws.cell(r, 22, value=low); low_n += 1

wb.save(FILE)
print(f'처리: {proc}, 전고: {high_n}, 전저: {low_n}, 실패: {len(fails)}')
for f in fails[:10]:
    print(f'  행{f[0]} {f[1]} {f[2]} {f[3]} 전용{f[4]}')
